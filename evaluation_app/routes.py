import json
import base64
import re
import uuid
from difflib import SequenceMatcher
from io import BytesIO
from datetime import datetime
from functools import wraps
from pathlib import Path

from flask import (
    Blueprint, Response, abort, flash, jsonify, redirect, render_template,
    current_app, request, send_file, send_from_directory, session, url_for,
)
from flask_login import current_user, login_required
from werkzeug.security import check_password_hash, generate_password_hash

from . import db
from .account_import import build_account_template, parse_account_workbook
from .dictionary_service import (
    blank_template, calculate_score, current_scheme, current_year, export_dictionary,
    full_example, import_dictionary, indicator_to_dict, validate_dictionary,
)
from .models import (
    AIModelSetting, AcademicYear, AuditLog, Category, EvaluationRecord,
    EvaluationScheme, Indicator, RecordAttachment, SchemeMembership, SystemSetting, User,
)

main_bp = Blueprint('main', __name__)


def admin_required(view):
    @wraps(view)
    @login_required
    def wrapped(*args, **kwargs):
        if current_user.role not in {'superadmin', 'admin'}:
            abort(403)
        return view(*args, **kwargs)
    return wrapped


def reviewer_required(view):
    @wraps(view)
    @login_required
    def wrapped(*args, **kwargs):
        if current_user.role not in {'superadmin', 'admin', 'reviewer'}:
            abort(403)
        return view(*args, **kwargs)
    return wrapped


def superadmin_required(view):
    @wraps(view)
    @login_required
    def wrapped(*args, **kwargs):
        if current_user.role != 'superadmin':
            abort(403)
        return view(*args, **kwargs)
    return wrapped


def _ensure_ai_settings():
    settings = {
        item.purpose: item
        for item in AIModelSetting.query.filter(
            AIModelSetting.purpose.in_(['reasoning', 'vision']),
        ).order_by(AIModelSetting.id).all()
    }
    changed = False
    for purpose in ('reasoning', 'vision'):
        if purpose not in settings:
            settings[purpose] = AIModelSetting(purpose=purpose)
            db.session.add(settings[purpose])
            changed = True
    if changed:
        db.session.commit()
    return settings


def _ai_setting_for(purpose):
    return AIModelSetting.query.filter_by(purpose=purpose, enabled=True).order_by(AIModelSetting.id).first()


def _ai_setting_ready(setting):
    if not setting or not setting.api_base or not setting.model_name:
        return False
    from .providers import needs_api_key
    return not needs_api_key(setting.provider) or bool(setting.api_key)


def _accessible_schemes():
    query = EvaluationScheme.query
    if current_user.role == 'superadmin':
        return query.order_by(EvaluationScheme.status, EvaluationScheme.id).all()
    member_ids = db.session.query(SchemeMembership.scheme_id).filter_by(user_id=current_user.id)
    if current_user.role == 'admin':
        query = query.filter(db.or_(EvaluationScheme.owner_user_id == current_user.id, EvaluationScheme.id.in_(member_ids)))
    else:
        query = query.filter(db.or_(EvaluationScheme.id.in_(member_ids), EvaluationScheme.id == current_user.scheme_id))
    return query.order_by(EvaluationScheme.status, EvaluationScheme.id).all()


def _can_access_scheme(scheme_id):
    return any(item.id == scheme_id for item in _accessible_schemes())


def _scheme_teachers(scheme):
    if not scheme:
        return []
    member_ids = db.session.query(SchemeMembership.user_id).filter_by(scheme_id=scheme.id)
    return User.query.filter(
        User.role == 'teacher', User.is_active_flag.is_(True),
        db.or_(User.scheme_id == scheme.id, User.id.in_(member_ids)),
    ).order_by(User.display_name).all()


def _membership_role_for(role):
    return 'owner' if role == 'admin' else ('reviewer' if role == 'reviewer' else 'participant')


def _selected_schemes_from_form(fallback_current=False):
    raw_ids = request.form.getlist('scheme_ids')
    if not raw_ids and request.form.get('scheme_id'):
        raw_ids = [request.form.get('scheme_id')]
    if not raw_ids and fallback_current:
        fallback = current_scheme()
        if fallback:
            return [fallback]
    scheme_ids = []
    for raw_id in raw_ids:
        try:
            scheme_id = int(raw_id)
        except (TypeError, ValueError):
            raise ValueError('考评方案选择无效。') from None
        if scheme_id not in scheme_ids:
            scheme_ids.append(scheme_id)
    if not scheme_ids:
        raise ValueError('请至少选择一套考评方案。')
    schemes = [db.session.get(EvaluationScheme, scheme_id) for scheme_id in scheme_ids]
    if any(scheme is None for scheme in schemes):
        raise ValueError('所选考评方案不存在或已被删除。')
    return schemes


def _sync_user_schemes(user, schemes, role):
    target_ids = {scheme.id for scheme in schemes}
    membership_role = _membership_role_for(role)
    existing = {membership.scheme_id: membership for membership in list(user.scheme_memberships)}

    for scheme_id, membership in existing.items():
        if scheme_id not in target_ids:
            if membership.scheme.owner_user_id == user.id:
                membership.scheme.owner_user_id = None
            db.session.delete(membership)
        else:
            membership.membership_role = membership_role

    for scheme in schemes:
        if scheme.id not in existing:
            db.session.add(SchemeMembership(
                scheme_id=scheme.id, user_id=user.id, membership_role=membership_role,
            ))
        if role == 'admin' and scheme.owner_user_id is None:
            scheme.owner_user_id = user.id

    if role != 'admin':
        for scheme in EvaluationScheme.query.filter_by(owner_user_id=user.id).all():
            scheme.owner_user_id = None

    if user.scheme_id not in target_ids:
        user.scheme_id = schemes[0].id


@main_bp.route('/')
@login_required
def dashboard():
    year = current_year()
    if current_user.role in {'superadmin', 'admin'}:
        year_id = year.id if year else -1
        scheme = current_scheme()
        stats = {
            'teachers': len(_scheme_teachers(scheme)),
            'indicators': Indicator.query.filter_by(academic_year_id=year_id).count(),
            'teacher_indicators': Indicator.query.filter_by(
                academic_year_id=year_id, data_source='teacher', enabled=True,
            ).count(),
            'admin_indicators': Indicator.query.filter_by(
                academic_year_id=year_id, data_source='admin', enabled=True,
            ).count(),
            'records': EvaluationRecord.query.filter(
                EvaluationRecord.academic_year_id == year_id,
                EvaluationRecord.status != 'voided',
            ).count(),
            'pending': EvaluationRecord.query.filter_by(academic_year_id=year_id, status='pending').count(),
        }
    elif current_user.role == 'reviewer':
        year_id = year.id if year else -1
        stats = {
            'pending': EvaluationRecord.query.filter_by(academic_year_id=year_id, status='pending').count(),
            'approved': EvaluationRecord.query.filter_by(academic_year_id=year_id, status='approved').count(),
            'rejected': EvaluationRecord.query.filter_by(academic_year_id=year_id, status='rejected').count(),
            'records': EvaluationRecord.query.filter(
                EvaluationRecord.academic_year_id == year_id,
                EvaluationRecord.source == 'teacher',
                EvaluationRecord.status != 'voided',
            ).count(),
        }
    else:
        records = EvaluationRecord.query.filter_by(
            academic_year_id=year.id if year else -1,
            target_user_id=current_user.id,
        ).all()
        stats = {
            'records': len(records),
            'pending': sum(record.status == 'pending' for record in records),
            'approved': sum(record.status == 'approved' for record in records),
            'score': sum(record.final_score or 0 for record in records if record.status == 'approved'),
        }
    return render_template('dashboard.html', year=year, stats=stats)


@main_bp.route('/profile')
@login_required
def profile_page():
    return render_template('profile.html')


@main_bp.route('/profile/password', methods=['POST'])
@login_required
def change_my_password():
    current_password = request.form.get('current_password', '')
    new_password = request.form.get('new_password', '')
    confirm_password = request.form.get('confirm_password', '')
    if not check_password_hash(current_user.password_hash, current_password):
        flash('当前密码不正确。', 'danger')
    elif len(new_password) < 8:
        flash('新密码至少需要8个字符。', 'danger')
    elif new_password != confirm_password:
        flash('两次输入的新密码不一致。', 'danger')
    elif check_password_hash(current_user.password_hash, new_password):
        flash('新密码不能与当前密码相同。', 'danger')
    else:
        current_user.password_hash = generate_password_hash(new_password)
        db.session.add(AuditLog(
            user_id=current_user.id,
            action='user.password_change',
            entity_type='user',
            entity_id=str(current_user.id),
            detail_json='{}',
        ))
        db.session.commit()
        flash('密码修改成功，下次登录请使用新密码。', 'success')
    return redirect(url_for('main.profile_page'))


@main_bp.route('/admin/settings/ai', methods=['GET', 'POST'])
@superadmin_required
def ai_settings_page():
    from .providers import provider_list
    providers = provider_list()
    provider_keys = {item['key'] for item in providers}
    settings = _ensure_ai_settings()
    if request.method == 'POST':
        changed_models = {}
        for purpose, setting in settings.items():
            provider = request.form.get(f'{purpose}_provider', 'custom')
            setting.provider = provider if provider in provider_keys else 'custom'
            setting.api_base = _normalize_api_base(request.form.get(f'{purpose}_api_base', ''))
            setting.model_name = request.form.get(f'{purpose}_model_name', '').strip()
            supplied_key = request.form.get(f'{purpose}_api_key', '').strip()
            if supplied_key:
                setting.api_key = supplied_key
            setting.enabled = request.form.get(f'{purpose}_enabled') == 'on'
            changed_models[purpose] = {
                'provider': setting.provider,
                'model_name': setting.model_name,
                'enabled': setting.enabled,
            }
        db.session.add(AuditLog(
            user_id=current_user.id,
            action='ai.settings_update',
            entity_type='system_setting',
            entity_id='ai_models',
            detail_json=json.dumps(changed_models, ensure_ascii=False),
        ))
        db.session.commit()
        flash('两类 AI 模型配置已保存，系统会按用途自动调用。', 'success')
        return redirect(url_for('main.ai_settings_page'))
    return render_template(
        'ai_settings.html', settings=settings,
        has_keys={purpose: bool(setting.api_key) for purpose, setting in settings.items()},
        providers=providers,
    )


@main_bp.route('/admin/settings/update', methods=['GET', 'POST'])
@main_bp.route('/admin/settings', methods=['GET', 'POST'])
@superadmin_required
def update_settings_page():
    from .changelog import CHANGELOG
    from .update_service import status
    setting = db.session.get(SystemSetting, 'update_base_url')
    if setting is None:
        setting = SystemSetting(key='update_base_url', value='')
        db.session.add(setting)
        db.session.commit()
    if request.method == 'POST':
        setting.value = request.form.get('update_base_url', '').strip().rstrip('/')
        db.session.commit()
        flash('自定义更新源已保存。' if setting.value else '已恢复使用官方更新源。', 'success')
        return redirect(url_for('main.update_settings_page'))
    return render_template(
        'update_settings.html',
        update_base_url=setting.value,
        using_official_update_source=not bool(setting.value),
        update_status=status(Path(current_app.root_path).parent),
        changelog=CHANGELOG,
    )


@main_bp.route('/admin/settings/logo', methods=['POST'])
@superadmin_required
def school_logo_settings():
    setting = db.session.get(SystemSetting, 'school_logo_path')
    if setting is None:
        setting = SystemSetting(key='school_logo_path', value='')
        db.session.add(setting)
    logo_path = Path(current_app.config['UPLOAD_FOLDER']) / 'system' / 'school-logo.png'
    action = request.form.get('action', 'upload')
    if action == 'reset':
        logo_path.unlink(missing_ok=True)
        setting.value = ''
        audit_action = 'branding.logo_reset'
        message = '已恢复 CampusMetric 默认 Logo。'
    else:
        upload = request.files.get('logo')
        if not upload or not upload.filename:
            flash('请选择一张 Logo 图片。', 'danger')
            return redirect(url_for('main.update_settings_page'))
        try:
            image = _load_supported_image(upload.stream)
            if image.width * image.height > 40_000_000:
                raise ValueError('Logo 图片像素过大，请压缩后重试。')
            image.thumbnail((512, 512))
            if image.mode not in {'RGB', 'RGBA'}:
                image = image.convert('RGBA' if 'transparency' in image.info else 'RGB')
            logo_path.parent.mkdir(parents=True, exist_ok=True)
            temporary = logo_path.with_name('.school-logo.tmp')
            image.save(temporary, 'PNG', optimize=True)
            temporary.replace(logo_path)
        except ValueError as error:
            flash(str(error).replace('附件', 'Logo'), 'danger')
            return redirect(url_for('main.update_settings_page'))
        except Exception:
            current_app.logger.exception('Unable to save school logo')
            flash('Logo 图片处理失败，请更换图片后重试。', 'danger')
            return redirect(url_for('main.update_settings_page'))
        setting.value = 'system/school-logo.png'
        audit_action = 'branding.logo_update'
        message = '学校 Logo 已更新，左侧栏和手机端将立即使用新图标。'
    db.session.add(AuditLog(
        user_id=current_user.id,
        action=audit_action,
        entity_type='system_setting',
        entity_id='school_logo_path',
        detail_json='{}',
    ))
    db.session.commit()
    flash(message, 'success')
    return redirect(url_for('main.update_settings_page'))


@main_bp.route('/branding/school-logo')
def school_logo_asset():
    setting = db.session.get(SystemSetting, 'school_logo_path')
    if not setting or setting.value != 'system/school-logo.png':
        abort(404)
    folder = Path(current_app.config['UPLOAD_FOLDER']) / 'system'
    if not (folder / 'school-logo.png').is_file():
        abort(404)
    return send_from_directory(folder, 'school-logo.png', mimetype='image/png', max_age=3600)


@main_bp.route('/api/admin/update/check', methods=['POST'])
@superadmin_required
def check_update_api():
    from .update_service import APP_VERSION, effective_update_base_url, fetch_manifest, version_tuple
    setting = db.session.get(SystemSetting, 'update_base_url')
    base_url = effective_update_base_url(setting.value if setting else '')
    try:
        manifest = fetch_manifest(base_url)
    except Exception:
        return jsonify({'ok': False, 'message': '无法连接更新源或清单格式不正确。'}), 502
    newer = version_tuple(manifest['version']) > version_tuple(APP_VERSION)
    session_key = db.session.get(SystemSetting, 'pending_update_manifest')
    if session_key is None:
        session_key = SystemSetting(key='pending_update_manifest')
        db.session.add(session_key)
    session_key.value = json.dumps(manifest, ensure_ascii=False)
    db.session.commit()
    return jsonify({'ok': True, 'newer': newer, 'current': APP_VERSION, 'latest': manifest['version'],
                    'notes': manifest.get('notes', ''), 'size': manifest.get('size', 0)})


@main_bp.route('/api/admin/update/download', methods=['POST'])
@superadmin_required
def download_update_api():
    from .update_service import download_and_stage, effective_update_base_url
    base = db.session.get(SystemSetting, 'update_base_url')
    pending = db.session.get(SystemSetting, 'pending_update_manifest')
    if not pending or not pending.value:
        return jsonify({'ok': False, 'message': '请先检查更新。'}), 400
    try:
        manifest = json.loads(pending.value)
        download_and_stage(
            Path(current_app.root_path).parent,
            effective_update_base_url(base.value if base else ''),
            manifest,
        )
    except Exception as error:
        return jsonify({'ok': False, 'message': str(error)[:160]}), 502
    return jsonify({'ok': True, 'message': '更新包已下载、校验并安全暂存。'})


@main_bp.route('/api/admin/update/apply', methods=['POST'])
@superadmin_required
def apply_update_api():
    import os
    import threading
    from .update_service import apply_staged_update
    if os.getenv('CE_MANAGED_LAUNCHER') != '1':
        return jsonify({'ok': False, 'message': '当前为源码开发模式：可检查、下载和校验，不自动覆盖源码。免安装包由启动器托管时才允许应用。'}), 409
    try:
        replaced, rollback = apply_staged_update(Path(current_app.root_path).parent)
    except Exception as error:
        return jsonify({'ok': False, 'message': str(error)[:160]}), 500
    sentinel = Path(current_app.root_path).parent / 'backups' / '.restart_pending'
    sentinel.parent.mkdir(parents=True, exist_ok=True)
    sentinel.write_text(datetime.now().isoformat(), encoding='utf-8')
    threading.Timer(1.2, lambda: os._exit(0)).start()
    return jsonify({'ok': True, 'message': f'已替换 {replaced} 个程序文件，旧版已备份到 {rollback.name}。程序即将自动重启。'})


@main_bp.route('/api/admin/ai/test', methods=['POST'])
@superadmin_required
def test_ai_api():
    payload = request.get_json(silent=True) or {}
    purpose = payload.get('purpose', 'reasoning')
    if purpose not in {'reasoning', 'vision'}:
        return jsonify({'ok': False, 'message': '模型用途无效。'}), 400
    setting = _ai_setting_for(purpose)
    purpose_label = '文本思考模型' if purpose == 'reasoning' else '视觉识别模型'
    if not _ai_setting_ready(setting):
        return jsonify({'ok': False, 'message': f'请先完整保存并启用{purpose_label}配置。'}), 400
    try:
        content = _call_openai(setting, [{
            'role': 'user', 'content': '请只回复“连接成功”。',
        }], max_tokens=20)
        return jsonify({'ok': True, 'message': content[:100]})
    except Exception:
        return jsonify({'ok': False, 'message': f'{purpose_label}连接失败，请检查 Base URL、Key 和模型 ID。'}), 502


@main_bp.route('/admin/dictionary')
@admin_required
def dictionary_page():
    dictionary = export_dictionary(include_ids=True)
    return render_template('dictionary.html', dictionary=dictionary, year=current_year())


@main_bp.route('/admin/dictionary/manual')
@admin_required
def dictionary_manual_page():
    year = current_year()
    categories = Category.query.filter_by(academic_year_id=year.id if year else -1).order_by(Category.sort_order, Category.id).all()
    locked_indicator_ids = {
        row[0] for row in db.session.query(EvaluationRecord.indicator_id).filter(
            EvaluationRecord.academic_year_id == (year.id if year else -1),
        ).distinct().all()
    }
    return render_template(
        'dictionary_manual.html', year=year, categories=categories,
        locked_indicator_ids=locked_indicator_ids,
    )


@main_bp.route('/admin/dictionary/categories', methods=['POST'])
@admin_required
def create_category():
    year = current_year()
    if not year:
        flash('当前没有可编辑的学年。', 'danger')
        return redirect(url_for('main.dictionary_manual_page'))
    code = request.form.get('code', '').strip().lower()
    name = request.form.get('name', '').strip()
    if not code or not name:
        flash('分类编码和名称不能为空。', 'danger')
    elif Category.query.filter_by(academic_year_id=year.id, code=code).first():
        flash('该分类编码已存在。', 'danger')
    else:
        category = Category(
            academic_year_id=year.id, code=code, name=name,
            description=request.form.get('description', '').strip(),
            sort_order=_int_form('sort_order', 0), max_score=_optional_float_form('max_score'),
        )
        db.session.add(category)
        db.session.commit()
        flash('一级指标已创建。', 'success')
    return redirect(url_for('main.dictionary_manual_page'))


@main_bp.route('/admin/dictionary/categories/<int:category_id>', methods=['POST'])
@admin_required
def update_category(category_id):
    category = db.get_or_404(Category, category_id)
    if not _can_access_scheme(category.academic_year.scheme_id):
        abort(403)
    action = request.form.get('action', 'save')
    if action == 'delete':
        if EvaluationRecord.query.join(Indicator).filter(Indicator.category_id == category.id).first():
            flash('该分类已有考评记录，不能删除。', 'danger')
        else:
            db.session.delete(category)
            db.session.commit()
            flash('一级指标已删除。', 'success')
        return redirect(url_for('main.dictionary_manual_page'))
    category.name = request.form.get('name', '').strip() or category.name
    if 'description' in request.form:
        category.description = request.form.get('description', '').strip()
    category.sort_order = _int_form('sort_order', category.sort_order)
    category.max_score = _optional_float_form('max_score')
    db.session.commit()
    flash('一级指标已保存。', 'success')
    return redirect(url_for('main.dictionary_manual_page'))


@main_bp.route('/admin/dictionary/indicators', methods=['POST'])
@admin_required
def create_indicator():
    year = current_year()
    category = db.session.get(Category, request.form.get('category_id'))
    if not year or not category or category.academic_year_id != year.id:
        abort(400)
    try:
        rule = json.loads(request.form.get('scoring_rule', '{}'))
        tracking = json.loads(request.form.get('secondary_tracking', '{"enabled": false}'))
    except json.JSONDecodeError:
        flash('计分规则或二级跟踪字段不是有效 JSON。', 'danger')
        return redirect(url_for('main.dictionary_manual_page'))
    item = _indicator_form_data(rule, tracking)
    validation = validate_dictionary({
        'schema_version': '1.0', 'dictionary_name': '手动配置校验', 'description': '',
        'categories': [{'code': category.code, 'name': category.name, 'description': '', 'order': 1,
                        'max_score': category.max_score, 'items': [item]}],
    })
    if not validation['valid']:
        flash('；'.join(issue['message'] for issue in validation['errors'][:3]), 'danger')
    elif Indicator.query.filter_by(academic_year_id=year.id, code=item['code']).first():
        flash('该指标编码已存在。', 'danger')
    else:
        indicator = Indicator(
            academic_year_id=year.id, category_id=category.id, code=item['code'], name=item['name'],
            description=item['description'], sort_order=item['order'], score_group=item['score_group'],
            data_source=item['data_source'], scoring_type=item['scoring_type'],
            scoring_rule_json=json.dumps(rule, ensure_ascii=False),
            allow_multiple_records=item['allow_multiple_records'], requires_evidence=item['requires_evidence'],
            ai_enabled=item['ai_enabled'], secondary_tracking_json=json.dumps(tracking, ensure_ascii=False),
            enabled=item['enabled'],
        )
        db.session.add(indicator)
        db.session.commit()
        flash('考评明细已创建。', 'success')
    return redirect(url_for('main.dictionary_manual_page'))


@main_bp.route('/admin/dictionary/indicators/<int:indicator_id>', methods=['POST'])
@admin_required
def update_indicator(indicator_id):
    indicator = db.get_or_404(Indicator, indicator_id)
    if not _can_access_scheme(indicator.academic_year.scheme_id):
        abort(403)
    if request.form.get('action') == 'delete':
        if EvaluationRecord.query.filter_by(indicator_id=indicator.id).first():
            flash('该指标已有考评记录，只能停用，不能删除。', 'danger')
        else:
            db.session.delete(indicator)
            db.session.commit()
            flash('考评明细已删除。', 'success')
        return redirect(url_for('main.dictionary_manual_page'))
    has_records = EvaluationRecord.query.filter_by(indicator_id=indicator.id).first() is not None
    if not has_records and 'scoring_rule' in request.form:
        try:
            rule = json.loads(request.form.get('scoring_rule', '{}'))
            tracking = json.loads(request.form.get('secondary_tracking', '{"enabled": false}'))
        except json.JSONDecodeError:
            flash('计分规则或二级跟踪字段不是有效 JSON。', 'danger')
            return redirect(url_for('main.dictionary_manual_page'))
        item = _indicator_form_data(rule, tracking)
        category = indicator.category
        validation = validate_dictionary({
            'schema_version': '1.0', 'dictionary_name': '手动配置校验', 'description': '',
            'categories': [{'code': category.code, 'name': category.name, 'description': '', 'order': 1,
                            'max_score': category.max_score, 'items': [item]}],
        })
        if not validation['valid']:
            flash('；'.join(issue['message'] for issue in validation['errors'][:3]), 'danger')
            return redirect(url_for('main.dictionary_manual_page'))
        indicator.data_source = item['data_source']
        indicator.score_group = item['score_group']
        indicator.scoring_type = item['scoring_type']
        indicator.scoring_rule_json = json.dumps(rule, ensure_ascii=False)
        indicator.allow_multiple_records = item['allow_multiple_records']
        indicator.secondary_tracking_json = json.dumps(tracking, ensure_ascii=False)
    indicator.name = request.form.get('name', '').strip() or indicator.name
    if 'description' in request.form:
        indicator.description = request.form.get('description', '').strip()
    indicator.sort_order = _int_form('sort_order', indicator.sort_order)
    indicator.enabled = request.form.get('enabled') == 'on'
    indicator.requires_evidence = request.form.get('requires_evidence') == 'on'
    indicator.ai_enabled = request.form.get('ai_enabled') == 'on' and indicator.requires_evidence
    db.session.commit()
    flash('考评明细已保存。' + ('该项已有数据，本次未修改计分规则。' if has_records else ''), 'success')
    return redirect(url_for('main.dictionary_manual_page'))


@main_bp.route('/admin/dictionary/download/<kind>')
@admin_required
def dictionary_download(kind):
    if kind == 'current':
        data = export_dictionary()
        if data is None:
            abort(404)
        scheme = current_scheme()
        base = (scheme.name.strip() if scheme and scheme.name and scheme.name.strip() else '考评字典')
        filename = f'{base}-考评字典-{datetime.now().strftime("%Y%m%d")}.json'
    else:
        choices = {
            'blank': ('考评字典空白模板.json', blank_template()),
            'example': ('考评字典完整示例.json', full_example()),
        }
        if kind not in choices or choices[kind][1] is None:
            abort(404)
        filename, data = choices[kind]
    return Response(
        json.dumps(data, ensure_ascii=False, indent=2),
        mimetype='application/json',
        headers={'Content-Disposition': f"attachment; filename*=UTF-8''{_quote(filename)}"},
    )


@main_bp.route('/admin/dictionary/guide')
@admin_required
def dictionary_guide():
    text = '''CampusMetric JSON 考评字典字段说明（模板版本 1.0）

顶层字段：schema_version、dictionary_name、description、categories。
score_group：base / bonus / deduction。
data_source：teacher / admin / excel。
scoring_type：manual_score / fixed_score / tiered_score / count_score /
count_deduction / range_score / matrix_score / tenure_score / fixed_bonus。

请同时下载“考评字典完整示例.json”，其中包含全部积分类型的 scoring_rule 写法。

给 AI 的建议提示词：
请根据我提供的学校考评方案，严格按照这份 JSON 模板生成考评字典。只能使用字段说明中允许的计分类型和数据来源，不要添加程序代码或自定义公式。原方案未明确的分数不要自行猜测，请在 description 中标记“待管理员确认”。输出必须是可以直接解析的标准 JSON，不要添加 Markdown 代码围栏或额外说明。
'''
    return Response(
        text,
        mimetype='text/plain',
        headers={'Content-Disposition': f"attachment; filename*=UTF-8''{_quote('JSON考评字典字段说明.txt')}"},
    )


@main_bp.route('/api/admin/dictionary/validate', methods=['POST'])
@admin_required
def dictionary_validate_api():
    data = request.get_json(silent=True)
    if data is None:
        return jsonify({'valid': False, 'errors': [{'path': '$', 'message': '无法解析JSON文件'}], 'warnings': [], 'summary': {}}), 400
    return jsonify(validate_dictionary(data))


@main_bp.route('/api/admin/dictionary/import', methods=['POST'])
@admin_required
def dictionary_import_api():
    payload = request.get_json(silent=True) or {}
    if payload.get('confirmed') is not True:
        return jsonify({'ok': False, 'message': '请先核对并确认导入内容'}), 400
    data = payload.get('dictionary')
    try:
        result = import_dictionary(data, current_user.id)
    except ValueError as error:
        return jsonify({'ok': False, 'message': str(error)}), 409
    if not result['valid']:
        return jsonify({'ok': False, **result}), 422
    db.session.add(AuditLog(
        user_id=current_user.id,
        action='dictionary.import',
        entity_type='academic_year',
        entity_id=str(current_year().id),
        detail_json=json.dumps(result['summary'], ensure_ascii=False),
    ))
    db.session.commit()
    return jsonify({'ok': True, **result})


@main_bp.route('/api/admin/dictionary/from-document', methods=['POST'])
@admin_required
def dictionary_from_document_api():
    """Extract PDF/Markdown, ask the configured model for schema JSON, validate, then import."""
    from pypdf import PdfReader

    upload = request.files.get('file')
    setting = _ai_setting_for('reasoning')
    if not upload or not upload.filename:
        return jsonify({'ok': False, 'message': '请选择 PDF 或 Markdown 评价方案。'}), 400
    if not _ai_setting_ready(setting):
        return jsonify({'ok': False, 'message': '超级管理员尚未完成可用的文本思考模型配置。'}), 409
    suffix = Path(upload.filename).suffix.lower()
    if suffix not in {'.pdf', '.md', '.markdown'}:
        return jsonify({'ok': False, 'message': '只支持 PDF、MD 或 Markdown 文件。'}), 415
    try:
        raw = upload.read()
        if suffix == '.pdf':
            reader = PdfReader(BytesIO(raw))
            text_content = '\n'.join((page.extract_text() or '') for page in reader.pages)
        else:
            text_content = raw.decode('utf-8-sig')
    except Exception:
        return jsonify({'ok': False, 'message': '无法读取方案文件；扫描版 PDF 请先转换为可复制文字的 PDF。'}), 400
    text_content = text_content.strip()
    if not text_content:
        return jsonify({'ok': False, 'message': '方案中没有提取到可识别文字。'}), 400
    if len(text_content) > 80000:
        text_content = text_content[:80000]
    template = full_example()
    prompt = (
        '你是学校评价量化配置专家。请将下方评价方案转换成 CampusMetric JSON。'
        '必须严格沿用示例的字段、枚举和数据结构；每一项材料或每一次获奖都应允许一条一条录入，'
        '需要材料的项目 requires_evidence=true 且 allow_multiple_records=true；'
        '不得臆造原文没有的分数，信息不足时写入 description 标记“待管理员确认”。'
        '只输出一个标准 JSON 对象，不要 Markdown 围栏或解释。\n\n'
        f'合法完整示例：\n{json.dumps(template, ensure_ascii=False)}\n\n'
        f'待转换评价方案：\n{text_content}'
    )
    try:
        content = _call_openai(setting, [{'role': 'user', 'content': prompt}], max_tokens=8000)
        generated = _parse_json_object(content)
        result = import_dictionary(generated, current_user.id)
    except ValueError as error:
        return jsonify({'ok': False, 'message': str(error)}), 422
    except Exception:
        return jsonify({'ok': False, 'message': 'AI 生成或解析失败，请检查模型能力后重试。'}), 502
    if not result['valid']:
        return jsonify({'ok': False, 'message': 'AI 生成结果未通过字典校验。', **result}), 422
    db.session.add(AuditLog(
        user_id=current_user.id, action='dictionary.ai_document_import', entity_type='academic_year',
        entity_id=str(current_year().id), detail_json=json.dumps({'filename': upload.filename, **result['summary']}, ensure_ascii=False),
    ))
    db.session.commit()
    return jsonify({'ok': True, 'message': '方案已由 AI 转换、校验并写入当前方案。', 'dictionary': generated, **result})


@main_bp.route('/api/entry-indicators')
@login_required
def entry_indicators_api():
    year = current_year()
    if not year:
        return jsonify({'categories': []})
    requested_mode = request.args.get('mode', 'teacher')
    if current_user.role in {'superadmin', 'admin'} and requested_mode == 'admin':
        allowed_sources = {'admin'}
    elif current_user.role in {'superadmin', 'admin'} and requested_mode == 'all':
        allowed_sources = {'admin', 'teacher'}
    else:
        allowed_sources = {'teacher'}
    indicators = Indicator.query.filter(
        Indicator.academic_year_id == year.id,
        Indicator.enabled.is_(True),
        Indicator.data_source.in_(allowed_sources),
    ).order_by(Indicator.category_id, Indicator.sort_order, Indicator.id).all()
    grouped = []
    by_category = {}
    for indicator in indicators:
        category = by_category.get(indicator.category_id)
        if category is None:
            category = {
                'id': indicator.category.id,
                'name': indicator.category.name,
                'max_score': indicator.category.max_score,
                'indicators': [],
            }
            by_category[indicator.category_id] = category
            grouped.append(category)
        category['indicators'].append(indicator_to_dict(indicator))
    return jsonify({'categories': grouped})


@main_bp.route('/entry/new')
@login_required
def teacher_entry_page():
    if current_user.role in {'superadmin', 'admin'}:
        return redirect(url_for('main.admin_entry_page'))
    return render_template('entry_form.html', mode='teacher', teachers=[], edit_record_id=None, edit_record_status=None)


@main_bp.route('/entry/<int:record_id>/edit')
@login_required
def teacher_edit_record_page(record_id):
    record = db.get_or_404(EvaluationRecord, record_id)
    if current_user.role != 'teacher' or record.source != 'teacher' or record.submitted_by_user_id != current_user.id:
        abort(403)
    if not _can_access_scheme(record.scheme_id):
        abort(403)
    if record.status not in {'pending', 'rejected'} or record.academic_year.status != 'active':
        flash('只有待审核或已退回的填报可以修改。', 'danger')
        return redirect(url_for('main.my_results_page'))
    session['active_scheme_id'] = record.scheme_id
    return render_template(
        'entry_form.html', mode='teacher', teachers=[], edit_record_id=record.id,
        edit_record_status=record.status,
    )


@main_bp.route('/admin/entry/new')
@admin_required
def admin_entry_page():
    scheme = current_scheme()
    teachers = _scheme_teachers(scheme)
    return render_template('entry_form.html', mode='admin', teachers=teachers, edit_record_id=None, edit_record_status=None)


@main_bp.route('/api/entry/recognize', methods=['POST'])
@login_required
def recognize_entry_image():
    """Recognize an evidence image before submission and map it to the selected indicator form."""
    year = current_year()
    indicator = db.session.get(Indicator, request.form.get('indicator_id', type=int))
    upload = request.files.get('file')
    if not year or not indicator or indicator.academic_year_id != year.id or not indicator.enabled:
        return jsonify({'ok': False, 'message': '当前考评指标不存在或已经停用。'}), 404
    if not _can_access_scheme(year.scheme_id):
        abort(403)
    if current_user.role not in {'superadmin', 'admin', 'teacher'}:
        abort(403)
    if current_user.role == 'teacher' and indicator.data_source != 'teacher':
        abort(403)
    if not indicator.requires_evidence or not indicator.ai_enabled:
        return jsonify({'ok': False, 'message': '该指标没有启用材料 AI 识别。'}), 409
    if not upload or not upload.filename:
        return jsonify({'ok': False, 'message': '请先选择一张证书或材料图片。'}), 400

    setting = _ai_setting_for('vision')
    if not _ai_setting_ready(setting):
        return jsonify({'ok': False, 'message': '超级管理员尚未完成可用的视觉识别模型配置。'}), 409
    try:
        image_url = _uploaded_image_data_url(upload)
    except ValueError as error:
        return jsonify({'ok': False, 'message': str(error)}), 400

    prompt = _entry_recognition_prompt(indicator)
    try:
        content = _call_openai(setting, [{'role': 'user', 'content': [
            {'type': 'text', 'text': prompt},
            {'type': 'image_url', 'image_url': {'url': image_url}},
        ]}], max_tokens=900)
        recognized = _parse_json_object(content)
        mapped = _sanitize_entry_recognition(indicator, recognized)
    except ValueError as error:
        return jsonify({'ok': False, 'message': str(error)}), 422
    except Exception:
        current_app.logger.exception('Entry evidence recognition failed')
        return jsonify({'ok': False, 'message': 'AI 识别失败，请检查视觉模型配置或稍后重试；也可以直接手动填写。'}), 502

    db.session.add(AuditLog(
        user_id=current_user.id, action='ai.entry_recognize', entity_type='indicator',
        entity_id=str(indicator.id), detail_json=json.dumps({
            'filename': Path(upload.filename).name[:255],
            'mapped_fields': mapped['filled_fields'],
        }, ensure_ascii=False),
    ))
    db.session.commit()
    return jsonify({
        'ok': True,
        'message': 'AI 识别完成，已自动填写：' + ('、'.join(mapped['filled_fields']) if mapped['filled_fields'] else '补充说明'),
        **mapped,
    })


@main_bp.route('/review')
@reviewer_required
def review_page():
    year = current_year()
    status = request.args.get('status', 'pending')
    if status not in {'pending', 'approved', 'rejected', 'all'}:
        status = 'pending'
    query = EvaluationRecord.query.filter_by(
        academic_year_id=year.id if year else -1,
        source='teacher',
    ).filter(EvaluationRecord.status != 'voided')
    if year:
        query = query.filter_by(scheme_id=year.scheme_id)
    if status != 'all':
        query = query.filter_by(status=status)
    records = query.order_by(EvaluationRecord.created_at.desc()).all()
    return render_template('review.html', records=records, rows=[_record_view(row) for row in records], status=status)


def _comparison_value(value):
    if isinstance(value, bool):
        return '是' if value else '否'
    if isinstance(value, (dict, list)):
        value = json.dumps(value, ensure_ascii=False, sort_keys=True)
    return re.sub(r'[\s\W_]+', '', str(value or '').lower(), flags=re.UNICODE)


def _record_inputs(record):
    try:
        data = json.loads(record.input_json or '{}')
    except json.JSONDecodeError:
        data = {}
    return data if isinstance(data, dict) else {}


def _duplicate_similarity(current, candidate):
    current_inputs = _record_inputs(current)
    candidate_inputs = _record_inputs(candidate)
    common_keys = sorted(set(current_inputs) & set(candidate_inputs))
    field_scores = []
    reasons = []
    current_labels = {item['key']: item['label'] for item in _record_input_fields(current, current_inputs)}
    for key in common_keys:
        left = _comparison_value(current_inputs.get(key))
        right = _comparison_value(candidate_inputs.get(key))
        if not left or not right:
            continue
        ratio = 1.0 if left == right else SequenceMatcher(None, left, right).ratio()
        field_scores.append(ratio)
        if ratio >= 0.92:
            reasons.append(f'{current_labels.get(key, key)}相同')

    input_score = sum(field_scores) / len(field_scores) if field_scores else 0.0
    generic_keys = {'qualified', 'count', 'score', 'value', 'years'}
    informative = any(key not in generic_keys for key in common_keys)
    input_weight = 0.70 if informative else 0.35

    tracking_left = _comparison_value(current.secondary_tracking_value)
    tracking_right = _comparison_value(candidate.secondary_tracking_value)
    tracking_score = 0.0
    if tracking_left and tracking_right:
        tracking_score = 1.0 if tracking_left == tracking_right else SequenceMatcher(
            None, tracking_left, tracking_right,
        ).ratio()
        if tracking_score >= 0.92:
            reasons.append('二级跟踪字段相同')

    note_left = _comparison_value(current.note)
    note_right = _comparison_value(candidate.note)
    note_score = 0.0
    if note_left and note_right:
        note_score = SequenceMatcher(None, note_left, note_right).ratio()
        if note_score >= 0.88:
            reasons.append('补充说明高度相似')

    teacher_score = 1.0 if current.target_user_id == candidate.target_user_id else 0.0
    if teacher_score:
        reasons.append('申报教师相同')
    score = (
        input_score * input_weight
        + tracking_score * 0.15
        + note_score * 0.10
        + teacher_score * 0.05
    )
    return min(score, 1.0), reasons


def _duplicate_record_payload(record):
    view = _record_view(record)
    return {
        'id': record.id,
        'teacher': view['teacher'],
        'indicator': view['indicator'],
        'category': view['category'],
        'academic_year': record.academic_year.label,
        'source': '管理员录入' if record.source == 'admin' else '教师填报',
        'status_label': view['status_label'],
        'input_fields': view['input_fields'],
        'tracking_label': view['tracking_label'],
        'tracking': view['tracking'] or '',
        'note': view['note'] or '',
        'score': record.final_score if record.final_score is not None else record.auto_score,
        'created_at': record.created_at.strftime('%Y-%m-%d %H:%M'),
        'attachments': [{
            'id': item.id,
            'name': item.original_name,
            'url': url_for('main.view_attachment', attachment_id=item.id),
        } for item in record.attachments],
    }


@main_bp.route('/api/review/<int:record_id>/duplicates')
@reviewer_required
def review_duplicate_check_api(record_id):
    record = db.get_or_404(EvaluationRecord, record_id)
    if not _can_access_scheme(record.scheme_id):
        abort(403)
    if record.source != 'teacher' or record.status == 'voided':
        return jsonify({'ok': False, 'message': '该记录不支持查重。'}), 409

    candidates = EvaluationRecord.query.filter(
        EvaluationRecord.id != record.id,
        EvaluationRecord.scheme_id == record.scheme_id,
        EvaluationRecord.indicator_id == record.indicator_id,
        EvaluationRecord.status == 'approved',
    ).order_by(EvaluationRecord.created_at.desc()).limit(150).all()
    matches = []
    for candidate in candidates:
        similarity, reasons = _duplicate_similarity(record, candidate)
        if similarity < 0.62:
            continue
        matches.append({
            'similarity': round(similarity * 100),
            'reasons': reasons[:5],
            'record': _duplicate_record_payload(candidate),
        })
    matches.sort(key=lambda item: (-item['similarity'], -item['record']['id']))
    matches = matches[:5]
    db.session.add(AuditLog(
        user_id=current_user.id,
        action='record.duplicate_check',
        entity_type='evaluation_record',
        entity_id=str(record.id),
        detail_json=json.dumps({
            'approved_candidates': len(candidates),
            'matches': [{'id': item['record']['id'], 'similarity': item['similarity']} for item in matches],
        }, ensure_ascii=False),
    ))
    db.session.commit()
    return jsonify({
        'ok': True,
        'message': f'发现 {len(matches)} 条相似的已通过记录。' if matches else '未发现相似的已通过记录。',
        'current': _duplicate_record_payload(record),
        'matches': matches,
    })


@main_bp.route('/review/<int:record_id>', methods=['POST'])
@reviewer_required
def review_record(record_id):
    record = db.get_or_404(EvaluationRecord, record_id)
    if not _can_access_scheme(record.scheme_id):
        abort(403)
    if record.academic_year.status != 'active':
        flash('已归档学年不能审核，请先还原。', 'danger')
        return redirect(url_for('main.review_page'))
    action = request.form.get('action')
    note = request.form.get('review_note', '').strip()
    if record.source != 'teacher':
        flash('该记录不需要审核。', 'danger')
        return redirect(url_for('main.review_page'))
    if action == 'void':
        if record.status != 'approved':
            flash('只能作废已审核通过的申报记录。', 'danger')
            return redirect(url_for('main.review_page', status='approved'))
        if not note:
            flash('作废已通过记录时请填写原因。', 'danger')
            return redirect(url_for('main.review_page', status='approved'))
        before_score = record.final_score
        record.status = 'voided'
        record.final_score = None
        record.review_note = f'作废原因：{note}'
        message = '已作废该条已通过申报，记录现仅对申请教师可见，由教师自行删除。'
        audit_action = 'record.voided_by_reviewer'
        audit_detail = {'reason': note, 'before_score': before_score}
    elif record.status != 'pending':
        flash('该记录已处理或不需要审核。', 'danger')
        return redirect(url_for('main.review_page'))
    elif action == 'approve':
        record.status = 'approved'
        record.final_score = record.auto_score
        message = '已通过该条填报。'
    elif action == 'reject':
        if not note:
            flash('退回时请填写原因。', 'danger')
            return redirect(url_for('main.review_page'))
        record.status = 'rejected'
        record.final_score = None
        message = '已退回该条填报。'
    else:
        abort(400)
    if action != 'void':
        record.review_note = note
        audit_action = f'record.{record.status}'
        audit_detail = {'review_note': note, 'final_score': record.final_score}
    record.reviewed_by_user_id = current_user.id
    record.reviewed_at = datetime.now()
    db.session.add(AuditLog(
        user_id=current_user.id,
        action=audit_action,
        entity_type='evaluation_record',
        entity_id=str(record.id),
        detail_json=json.dumps(audit_detail, ensure_ascii=False),
    ))
    db.session.commit()
    flash(message, 'success')
    return redirect(url_for('main.review_page'))


@main_bp.route('/my/records/<int:record_id>/void', methods=['POST'])
@login_required
def void_my_record(record_id):
    record = db.get_or_404(EvaluationRecord, record_id)
    if current_user.role != 'teacher' or record.source != 'teacher' or record.submitted_by_user_id != current_user.id:
        abort(403)
    if not _can_access_scheme(record.scheme_id):
        abort(403)
    if record.academic_year.status != 'active':
        flash('已归档学年不能作废申报，请先由管理员还原。', 'danger')
    elif record.status not in {'pending', 'rejected'}:
        flash('只有待审核或已退回的申报可由教师作废。', 'danger')
    else:
        previous_status = record.status
        record.status = 'voided'
        record.final_score = None
        record.review_note = '申请人已主动作废该条申报。'
        db.session.add(AuditLog(
            user_id=current_user.id,
            action='record.voided_by_teacher',
            entity_type='evaluation_record',
            entity_id=str(record.id),
            detail_json=json.dumps({'before_status': previous_status}, ensure_ascii=False),
        ))
        db.session.commit()
        flash('申报已作废，现在可以由你自行删除。', 'success')
    return redirect(url_for('main.my_results_page'))


@main_bp.route('/my/records/<int:record_id>/delete', methods=['POST'])
@login_required
def delete_my_voided_record(record_id):
    record = db.get_or_404(EvaluationRecord, record_id)
    if current_user.role != 'teacher' or record.source != 'teacher' or record.submitted_by_user_id != current_user.id:
        abort(403)
    if not _can_access_scheme(record.scheme_id):
        abort(403)
    if record.academic_year.status != 'active':
        flash('已归档学年不能删除记录，请先由管理员还原。', 'danger')
        return redirect(url_for('main.my_results_page'))
    if record.status != 'voided':
        flash('记录必须先作废，才能删除。', 'danger')
        return redirect(url_for('main.my_results_page'))

    attachment_paths = [Path(current_app.config['UPLOAD_FOLDER']) / item.stored_path for item in record.attachments]
    snapshot = {
        'scheme_id': record.scheme_id,
        'academic_year_id': record.academic_year_id,
        'indicator_id': record.indicator_id,
        'input_json': record.input_json,
        'secondary_tracking_value': record.secondary_tracking_value,
        'note': record.note,
        'attachment_names': [item.original_name for item in record.attachments],
    }
    deleted_id = record.id
    db.session.delete(record)
    db.session.add(AuditLog(
        user_id=current_user.id,
        action='record.deleted_by_teacher',
        entity_type='evaluation_record',
        entity_id=str(deleted_id),
        detail_json=json.dumps(snapshot, ensure_ascii=False),
    ))
    db.session.commit()
    for path in attachment_paths:
        try:
            path.unlink(missing_ok=True)
            path.parent.rmdir()
        except OSError:
            current_app.logger.warning('Unable to remove deleted record attachment path %s', path)
    flash('已删除作废申报及其材料。', 'success')
    return redirect(url_for('main.my_results_page'))


@main_bp.route('/my/results')
@login_required
def my_results_page():
    if current_user.role in {'superadmin', 'admin'}:
        return redirect(url_for('main.dashboard'))
    year = current_year()
    records = EvaluationRecord.query.filter_by(
        academic_year_id=year.id if year else -1,
        target_user_id=current_user.id,
    ).order_by(EvaluationRecord.created_at.desc()).all()
    rows = [_record_view(row) for row in records]
    totals = {
        'score': sum(row.final_score or 0 for row in records if row.status == 'approved'),
        'approved': sum(row.status == 'approved' for row in records),
        'pending': sum(row.status == 'pending' for row in records),
        'rejected': sum(row.status == 'rejected' for row in records),
        'voided': sum(row.status == 'voided' for row in records),
    }
    return render_template('results.html', rows=rows, totals=totals, year=year)


@main_bp.route('/admin/users', methods=['GET', 'POST'])
@superadmin_required
def users_page():
    if request.method == 'POST':
        username = request.form.get('username', '').strip().lower()
        display_name = request.form.get('display_name', '').strip()
        password = request.form.get('password', '')
        role = request.form.get('role', 'teacher')
        employee_no = request.form.get('employee_no', '').strip() or None
        try:
            selected_schemes = _selected_schemes_from_form(fallback_current=True)
            scheme_error = None
        except ValueError as error:
            selected_schemes = []
            scheme_error = str(error)
        if role not in {'admin', 'reviewer', 'teacher'}:
            flash('账号角色无效。', 'danger')
        elif scheme_error:
            flash(scheme_error, 'danger')
        elif not username or not display_name:
            flash('请填写账号和姓名。', 'danger')
        elif len(password) < 8:
            flash('初始密码至少需要8个字符。', 'danger')
        elif User.query.filter_by(username=username).first():
            flash('该登录账号已存在。', 'danger')
        elif employee_no and User.query.filter_by(employee_no=employee_no).first():
            flash('该工号已存在。', 'danger')
        else:
            user = User(
                username=username,
                display_name=display_name,
                password_hash=generate_password_hash(password),
                role=role,
                employee_no=employee_no,
                scheme_id=selected_schemes[0].id,
                is_homeroom_teacher=role == 'teacher' and request.form.get('is_homeroom_teacher') == 'on',
                is_grade_leader=role == 'teacher' and request.form.get('is_grade_leader') == 'on',
            )
            db.session.add(user)
            db.session.flush()
            _sync_user_schemes(user, selected_schemes, role)
            db.session.add(AuditLog(
                user_id=current_user.id, action='user.create', entity_type='user', entity_id=str(user.id),
                detail_json=json.dumps({
                    'username': username, 'role': role,
                    'scheme_ids': [scheme.id for scheme in selected_schemes],
                }, ensure_ascii=False),
            ))
            db.session.commit()
            flash(f'已创建账号：{display_name}', 'success')
            return redirect(url_for('main.users_page'))
    users = User.query.order_by(User.role, User.display_name).all()
    schemes = EvaluationScheme.query.order_by(EvaluationScheme.id).all()
    return render_template('users.html', users=users, schemes=schemes)


@main_bp.route('/admin/users/import-template')
@superadmin_required
def account_import_template():
    schemes = EvaluationScheme.query.order_by(EvaluationScheme.code).all()
    output = build_account_template([scheme.code for scheme in schemes])
    return send_file(
        output, as_attachment=True, download_name='CampusMetric账号导入模板.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@main_bp.route('/admin/users/import', methods=['POST'])
@superadmin_required
def import_accounts():
    upload = request.files.get('file')
    if not upload or not upload.filename:
        flash('请选择要导入的 Excel 文件。', 'danger')
        return redirect(url_for('main.users_page'))
    if not upload.filename.lower().endswith('.xlsx'):
        flash('账号导入仅支持系统模板格式的 .xlsx 文件。', 'danger')
        return redirect(url_for('main.users_page'))

    schemes = EvaluationScheme.query.order_by(EvaluationScheme.id).all()
    records, errors = parse_account_workbook(
        upload, schemes,
        {username for username, in db.session.query(User.username).all()},
        {employee_no for employee_no, in db.session.query(User.employee_no).filter(User.employee_no.is_not(None)).all()},
    )
    if errors:
        shown = errors[:8]
        suffix = f'；另有 {len(errors) - len(shown)} 个问题' if len(errors) > len(shown) else ''
        flash('导入未写入任何账号：' + '；'.join(shown) + suffix, 'danger')
        return redirect(url_for('main.users_page'))

    scheme_by_code = {scheme.code: scheme for scheme in schemes}
    imported = []
    try:
        for item in records:
            selected_schemes = [scheme_by_code[code] for code in item['scheme_codes']]
            user = User(
                username=item['username'], display_name=item['display_name'], employee_no=item['employee_no'],
                password_hash=generate_password_hash(item['password']), role=item['role'],
                scheme_id=selected_schemes[0].id,
                is_homeroom_teacher=item['is_homeroom_teacher'], is_grade_leader=item['is_grade_leader'],
                tenure_years=item['tenure_years'], is_active_flag=item['is_active_flag'],
            )
            db.session.add(user)
            db.session.flush()
            _sync_user_schemes(user, selected_schemes, user.role)
            imported.append(user.username)
        db.session.add(AuditLog(
            user_id=current_user.id, action='user.bulk_import', entity_type='user', entity_id='',
            detail_json=json.dumps({
                'count': len(imported), 'usernames': imported,
                'source_filename': Path(upload.filename).name,
            }, ensure_ascii=False),
        ))
        db.session.commit()
    except Exception:
        db.session.rollback()
        current_app.logger.exception('Account bulk import failed')
        flash('账号导入时发生错误，系统已回滚，本次没有写入任何账号。', 'danger')
        return redirect(url_for('main.users_page'))

    flash(f'成功导入 {len(imported)} 个账号，账号已按方案建立关联。', 'success')
    return redirect(url_for('main.users_page'))


@main_bp.route('/admin/users/<int:user_id>/profile', methods=['POST'])
@superadmin_required
def update_user_profile(user_id):
    user = db.get_or_404(User, user_id)
    if user.role == 'superadmin':
        abort(403)
    role = request.form.get('role', user.role)
    if role not in {'admin', 'reviewer', 'teacher'}:
        abort(400)
    try:
        selected_schemes = _selected_schemes_from_form()
    except ValueError as error:
        flash(str(error), 'danger')
        return redirect(url_for('main.users_page'))
    before_scheme_ids = sorted({membership.scheme_id for membership in user.scheme_memberships} | ({user.scheme_id} if user.scheme_id else set()))
    before = {
        'role': user.role, 'is_homeroom_teacher': user.is_homeroom_teacher,
        'is_grade_leader': user.is_grade_leader, 'tenure_years': user.tenure_years,
        'scheme_id': user.scheme_id, 'scheme_ids': before_scheme_ids,
    }
    user.role = role
    user.is_homeroom_teacher = role == 'teacher' and request.form.get('is_homeroom_teacher') == 'on'
    user.is_grade_leader = role == 'teacher' and request.form.get('is_grade_leader') == 'on'
    try:
        user.tenure_years = max(0, int(request.form.get('tenure_years', 0)))
    except ValueError:
        user.tenure_years = 0
    _sync_user_schemes(user, selected_schemes, role)
    after = {
        'role': user.role, 'is_homeroom_teacher': user.is_homeroom_teacher,
        'is_grade_leader': user.is_grade_leader, 'tenure_years': user.tenure_years,
        'scheme_id': user.scheme_id, 'scheme_ids': [scheme.id for scheme in selected_schemes],
    }
    db.session.add(AuditLog(
        user_id=current_user.id, action='user.profile_update', entity_type='user', entity_id=str(user.id),
        detail_json=json.dumps({'before': before, 'after': after}, ensure_ascii=False),
    ))
    db.session.commit()
    flash(f'已更新 {user.display_name} 的角色、标记与可用方案。', 'success')
    return redirect(url_for('main.users_page'))


@main_bp.route('/admin/users/<int:user_id>/toggle', methods=['POST'])
@superadmin_required
def toggle_user(user_id):
    user = db.get_or_404(User, user_id)
    if user.id == current_user.id:
        flash('不能停用当前登录的管理员账号。', 'danger')
        return redirect(url_for('main.users_page'))
    user.is_active_flag = not user.is_active_flag
    db.session.add(AuditLog(
        user_id=current_user.id, action='user.toggle', entity_type='user', entity_id=str(user.id),
        detail_json=json.dumps({'active': user.is_active_flag}, ensure_ascii=False),
    ))
    db.session.commit()
    flash(f'{user.display_name}已{"启用" if user.is_active_flag else "停用"}。', 'success')
    return redirect(url_for('main.users_page'))


@main_bp.route('/admin/users/<int:user_id>/password', methods=['POST'])
@superadmin_required
def reset_user_password(user_id):
    user = db.get_or_404(User, user_id)
    password = request.form.get('password', '')
    if len(password) < 8:
        flash('新密码至少需要8个字符。', 'danger')
    else:
        user.password_hash = generate_password_hash(password)
        db.session.add(AuditLog(
            user_id=current_user.id, action='user.password_reset', entity_type='user', entity_id=str(user.id),
            detail_json='{}',
        ))
        db.session.commit()
        flash(f'已重置 {user.display_name} 的密码。', 'success')
    return redirect(url_for('main.users_page'))


@main_bp.route('/admin/records')
@admin_required
def records_page():
    year = _selected_year()
    records = EvaluationRecord.query.filter_by(
        academic_year_id=year.id if year else -1,
    ).filter(EvaluationRecord.status != 'voided').order_by(EvaluationRecord.created_at.desc()).all()
    rows = [_record_view(row) for row in records]
    total = sum(record.final_score or 0 for record in records if record.status == 'approved')
    return render_template('records.html', rows=rows, total=total, year=year)


@main_bp.route('/admin/records/<int:record_id>/score', methods=['POST'])
@admin_required
def override_record_score(record_id):
    record = db.get_or_404(EvaluationRecord, record_id)
    if not _can_access_scheme(record.scheme_id):
        abort(403)
    if record.status == 'voided':
        abort(404)
    if record.academic_year.status != 'active':
        flash('已归档学年不能改分，请先还原。', 'danger')
        return redirect(url_for('main.records_page', year_id=record.academic_year_id))
    try:
        final_score = float(request.form.get('final_score', ''))
    except ValueError:
        flash('请填写有效分数。', 'danger')
        return redirect(url_for('main.records_page'))
    before = record.final_score
    reason = request.form.get('reason', '').strip()
    record.final_score = final_score
    record.status = 'approved'
    record.admin_overridden = record.auto_score is None or abs(final_score - record.auto_score) > 1e-9
    record.review_note = reason
    record.reviewed_by_user_id = current_user.id
    record.reviewed_at = datetime.now()
    db.session.add(AuditLog(
        user_id=current_user.id, action='record.score_override', entity_type='evaluation_record', entity_id=str(record.id),
        detail_json=json.dumps({'before': before, 'after': final_score, 'reason': reason}, ensure_ascii=False),
    ))
    db.session.commit()
    flash('分数已保存并直接生效。', 'success')
    return redirect(url_for('main.records_page'))


@main_bp.route('/admin/ranking')
@admin_required
def ranking_page():
    year = _selected_year()
    scheme = current_scheme()
    teachers = _scheme_teachers(scheme)
    rows = []
    for teacher in teachers:
        records = EvaluationRecord.query.filter_by(
            academic_year_id=year.id if year else -1, target_user_id=teacher.id, status='approved',
        ).all()
        rows.append({
            'teacher': teacher, 'score': sum(record.final_score or 0 for record in records),
            'records': len(records),
        })
    rows.sort(key=lambda row: (-row['score'], row['teacher'].display_name))
    last_score = None
    rank = 0
    for index, row in enumerate(rows, 1):
        if row['score'] != last_score:
            rank = index
            last_score = row['score']
        row['rank'] = rank
    return render_template('ranking.html', rows=rows, year=year, years=_scheme_years())


@main_bp.route('/admin/archive-analytics')
@admin_required
def archive_analytics_page():
    schemes = _accessible_schemes()
    scheme_ids = [item.id for item in schemes]
    teachers = _archive_teachers(scheme_ids)
    class_indicator_ids = _class_tracking_indicator_ids(scheme_ids)
    tracking_values = [row[0] for row in db.session.query(EvaluationRecord.secondary_tracking_value).filter(
        EvaluationRecord.scheme_id.in_(scheme_ids or [-1]),
        EvaluationRecord.indicator_id.in_(class_indicator_ids or [-1]),
        EvaluationRecord.status != 'voided',
        EvaluationRecord.secondary_tracking_value.isnot(None),
        EvaluationRecord.secondary_tracking_value != '',
    ).distinct().order_by(EvaluationRecord.secondary_tracking_value).all()]
    return render_template('archive_analytics.html', teachers=teachers, tracking_values=tracking_values,
                           schemes=schemes)


@main_bp.route('/api/admin/archive-analytics')
@admin_required
def archive_analytics_api():
    from collections import defaultdict
    from datetime import datetime as dt

    schemes = _accessible_schemes()
    scheme_ids = [item.id for item in schemes]
    if not scheme_ids:
        return jsonify({'ok': True, 'timeline': [], 'comparison': [], 'categories': [], 'summary': {}})
    entity_type = request.args.get('entity_type', 'teacher')
    entity_value = request.args.get('entity_value', '').strip()
    query = EvaluationRecord.query.filter(
        EvaluationRecord.scheme_id.in_(scheme_ids), EvaluationRecord.status == 'approved',
    )
    start = request.args.get('start', '').strip()
    end = request.args.get('end', '').strip()
    try:
        if start:
            query = query.filter(EvaluationRecord.created_at >= dt.strptime(start, '%Y-%m-%d'))
        if end:
            query = query.filter(EvaluationRecord.created_at < dt.strptime(end, '%Y-%m-%d').replace(hour=23, minute=59, second=59))
    except ValueError:
        return jsonify({'ok': False, 'message': '日期格式无效。'}), 400
    all_records = query.order_by(EvaluationRecord.created_at).all()
    if entity_type == 'class':
        class_ids = set(_class_tracking_indicator_ids(scheme_ids))
        class_records = [row for row in all_records if row.indicator_id in class_ids]
        selected = [row for row in class_records if (row.secondary_tracking_value or '') == entity_value]
        labels = sorted({row.secondary_tracking_value for row in class_records if row.secondary_tracking_value})
        comparison = [(label, sum((row.final_score or 0) for row in class_records if row.secondary_tracking_value == label)) for label in labels]
        selected_label = entity_value or '未选择班级'
    else:
        try:
            teacher_id = int(entity_value)
        except (TypeError, ValueError):
            teacher_id = 0
        selected = [row for row in all_records if row.target_user_id == teacher_id]
        teachers = _archive_teachers(scheme_ids)
        comparison = [(teacher.display_name, sum((row.final_score or 0) for row in all_records if row.target_user_id == teacher.id)) for teacher in teachers]
        teacher = db.session.get(User, teacher_id)
        selected_label = teacher.display_name if teacher else '未选择教师'
    monthly, categories = defaultdict(float), defaultdict(float)
    for row in selected:
        monthly[row.created_at.strftime('%Y-%m')] += row.final_score or 0
        categories[row.indicator.category.name] += row.final_score or 0
    comparison.sort(key=lambda item: (-item[1], item[0]))
    details = []
    for row in reversed(selected):
        view = _record_view(row)
        details.append({
            'id': row.id,
            'scheme': row.scheme.name if row.scheme else '未标注方案',
            'scheme_code': row.scheme.code if row.scheme else '',
            'year': row.academic_year.label,
            'teacher': view['teacher'], 'category': view['category'], 'indicator': view['indicator'],
            'score': row.final_score or 0, 'created_at': row.created_at.strftime('%Y-%m-%d %H:%M'),
            'input_fields': view['input_fields'], 'tracking_label': view['tracking_label'],
            'tracking': view['tracking'], 'note': view['note'],
            'attachments': [{
                'name': item.original_name,
                'url': url_for('main.view_attachment', attachment_id=item.id),
            } for item in row.attachments],
        })
    return jsonify({
        'ok': True,
        'timeline': [{'label': key, 'score': round(value, 2)} for key, value in sorted(monthly.items())],
        'comparison': [{'label': key, 'score': round(value, 2)} for key, value in comparison[:20]],
        'categories': [{'label': key, 'score': round(value, 2)} for key, value in sorted(categories.items(), key=lambda item: -item[1])],
        'details': details,
        'summary': {
            'label': selected_label, 'score': round(sum((row.final_score or 0) for row in selected), 2),
            'records': len(selected), 'years': len({row.academic_year_id for row in selected}),
            'schemes': len({row.scheme_id for row in selected}),
            'archived_records': sum(row.academic_year.status == 'archived' for row in selected),
        },
    })


def _archive_teachers(scheme_ids):
    member_ids = db.session.query(SchemeMembership.user_id).filter(
        SchemeMembership.scheme_id.in_(scheme_ids or [-1]),
    )
    return User.query.filter(
        User.role == 'teacher', User.is_active_flag.is_(True),
        db.or_(User.scheme_id.in_(scheme_ids or [-1]), User.id.in_(member_ids)),
    ).order_by(User.display_name, User.id).all()


def _class_tracking_indicator_ids(scheme_ids):
    year_ids = [row[0] for row in db.session.query(AcademicYear.id).filter(
        AcademicYear.scheme_id.in_(scheme_ids or [-1]),
    ).all()]
    indicators = Indicator.query.filter(Indicator.academic_year_id.in_(year_ids or [-1])).all()
    result = []
    for indicator in indicators:
        try:
            config = json.loads(indicator.secondary_tracking_json or '{}')
        except json.JSONDecodeError:
            continue
        if config.get('enabled') and '班级' in str(config.get('label') or ''):
            result.append(indicator.id)
    return result


@main_bp.route('/admin/results.xlsx')
@admin_required
def export_results_excel():
    from openpyxl import Workbook
    year = _selected_year()
    workbook = Workbook()
    summary = workbook.active
    summary.title = '总分排名'
    summary.append(['排名', '姓名', '工号', '总分', '班主任', '年段长'])
    scheme = current_scheme()
    teachers = _scheme_teachers(scheme)
    ranked = []
    for teacher in teachers:
        records = EvaluationRecord.query.filter_by(academic_year_id=year.id, target_user_id=teacher.id, status='approved').all()
        ranked.append((teacher, sum(record.final_score or 0 for record in records)))
    ranked.sort(key=lambda item: (-item[1], item[0].display_name))
    for index, (teacher, score) in enumerate(ranked, 1):
        summary.append([index, teacher.display_name, teacher.employee_no or '', score, '是' if teacher.is_homeroom_teacher else '否', '是' if teacher.is_grade_leader else '否'])
    detail = workbook.create_sheet('考评明细')
    detail.append(['教师', '一级指标', '考评项目', '来源', '状态', '自动分', '最终分', '说明', '时间'])
    for record in EvaluationRecord.query.filter(
        EvaluationRecord.academic_year_id == year.id,
        EvaluationRecord.status != 'voided',
    ).order_by(EvaluationRecord.id).all():
        detail.append([record.target_user.display_name, record.indicator.category.name, record.indicator.name,
                       record.source, record.status, record.auto_score, record.final_score, record.note,
                       record.created_at.strftime('%Y-%m-%d %H:%M')])
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name=f'{year.name}校园评价结果.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@main_bp.route('/admin/external-template.xlsx')
@admin_required
def external_import_template():
    from openpyxl import Workbook
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = '外部结果导入'
    sheet.append(['username', 'indicator_code', 'value', 'final_score', 'tracking', 'note'])
    sheet.append(['teacher01', 'student_satisfaction', 95, '', '三年1班', '学生满意率'])
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name='外部考评结果导入模板.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@main_bp.route('/admin/external-import', methods=['POST'])
@admin_required
def external_import():
    from openpyxl import load_workbook
    year = current_year()
    upload = request.files.get('file')
    if not year or not upload:
        flash('请选择要导入的 Excel 文件。', 'danger')
        return redirect(url_for('main.ranking_page'))
    try:
        workbook = load_workbook(upload, read_only=True, data_only=True)
        rows = list(workbook.active.iter_rows(values_only=True))
    except Exception:
        flash('无法读取 Excel 文件。', 'danger')
        return redirect(url_for('main.ranking_page'))
    headers = [str(value or '').strip() for value in rows[0]] if rows else []
    required = {'username', 'indicator_code'}
    if not required.issubset(headers):
        flash('Excel 缺少 username 或 indicator_code 列。', 'danger')
        return redirect(url_for('main.ranking_page'))
    created, errors = 0, []
    for number, values in enumerate(rows[1:], 2):
        data = dict(zip(headers, values))
        user = User.query.filter_by(username=str(data.get('username') or '').strip(), role='teacher').first()
        indicator = Indicator.query.filter_by(academic_year_id=year.id, code=str(data.get('indicator_code') or '').strip()).first()
        if not user or not indicator or indicator.data_source != 'excel':
            errors.append(f'第{number}行账号或外部指标不存在')
            continue
        inputs = {'value': data.get('value')}
        try:
            auto_score = calculate_score(indicator, inputs)
            final_score = float(data['final_score']) if data.get('final_score') not in (None, '') else auto_score
        except (ValueError, TypeError, KeyError):
            errors.append(f'第{number}行分数无效')
            continue
        db.session.add(EvaluationRecord(
            scheme_id=year.scheme_id, academic_year_id=year.id, indicator_id=indicator.id,
            target_user_id=user.id, submitted_by_user_id=current_user.id, source='excel', status='approved',
            input_json=json.dumps(inputs, ensure_ascii=False), secondary_tracking_value=str(data.get('tracking') or ''),
            auto_score=auto_score, final_score=final_score, note=str(data.get('note') or ''),
        ))
        created += 1
    db.session.commit()
    flash(f'已导入 {created} 条外部结果' + (f'，{len(errors)} 条失败' if errors else ''), 'success' if created else 'danger')
    return redirect(url_for('main.ranking_page'))


@main_bp.route('/admin/settings/scheme', methods=['GET', 'POST'])
@admin_required
def scheme_settings_page():
    scheme = current_scheme()
    if request.method == 'POST':
        action = request.form.get('action', 'save')
        code = request.form.get('code', '').strip().upper()
        name = request.form.get('name', '').strip()
        if not code or not name:
            flash('方案编号和名称不能为空。', 'danger')
        elif EvaluationScheme.query.filter(
            EvaluationScheme.code == code,
            EvaluationScheme.id != (scheme.id if scheme and action != 'create' else -1),
        ).first():
            flash('方案编号已存在。', 'danger')
        elif action == 'create':
            owner_id = current_user.id if current_user.role == 'admin' else None
            new_scheme = EvaluationScheme(
                code=code, name=name, description=request.form.get('description', '').strip(),
                owner_user_id=owner_id, status='active',
            )
            db.session.add(new_scheme)
            db.session.flush()
            if current_user.role == 'admin':
                db.session.add(SchemeMembership(scheme_id=new_scheme.id, user_id=current_user.id, membership_role='owner'))
            year_label = request.form.get('academic_year', '').strip() or '当前学年'
            internal_name = _unique_year_name(year_label, code)
            db.session.add(AcademicYear(
                name=internal_name, display_name=year_label, scheme_id=new_scheme.id, status='active',
            ))
            db.session.commit()
            session['active_scheme_id'] = new_scheme.id
            flash('新考评方案已创建并切换。', 'success')
            return redirect(url_for('main.scheme_settings_page'))
        else:
            scheme.code = code
            scheme.name = name
            scheme.description = request.form.get('description', '').strip()
            db.session.commit()
            flash('考评方案信息已保存。', 'success')
    return render_template('scheme_settings.html', scheme=scheme, years=_scheme_years(), schemes=_accessible_schemes())


@main_bp.route('/scheme/select/<int:scheme_id>', methods=['POST'])
@login_required
def select_scheme(scheme_id):
    if not _can_access_scheme(scheme_id):
        abort(403)
    session['active_scheme_id'] = scheme_id
    next_url = request.form.get('next', '')
    if not next_url.startswith('/') or next_url.startswith('//'):
        next_url = url_for('main.dashboard')
    return redirect(next_url)


@main_bp.route('/admin/schemes/<int:scheme_id>/archive', methods=['POST'])
@admin_required
def archive_scheme(scheme_id):
    scheme = db.get_or_404(EvaluationScheme, scheme_id)
    if not _can_access_scheme(scheme.id):
        abort(403)
    scheme.status = 'archived'
    AcademicYear.query.filter_by(scheme_id=scheme.id, status='active').update({
        'status': 'archived', 'archived_at': datetime.now(),
    })
    db.session.commit()
    session.pop('active_scheme_id', None)
    flash(f'{scheme.name}已归档，方案和学年数据均转为只读。', 'success')
    return redirect(url_for('main.scheme_settings_page'))


@main_bp.route('/admin/schemes/<int:scheme_id>/restore', methods=['POST'])
@admin_required
def restore_scheme(scheme_id):
    scheme = db.get_or_404(EvaluationScheme, scheme_id)
    if not _can_access_scheme(scheme.id):
        abort(403)
    scheme.status = 'active'
    db.session.commit()
    session['active_scheme_id'] = scheme.id
    flash(f'{scheme.name}已恢复，可选择学年继续使用。', 'success')
    return redirect(url_for('main.scheme_settings_page'))


@main_bp.route('/admin/years/<int:year_id>/archive', methods=['POST'])
@admin_required
def archive_year(year_id):
    year = db.get_or_404(AcademicYear, year_id)
    if not _can_access_scheme(year.scheme_id):
        abort(403)
    year.status = 'archived'
    year.archived_at = datetime.now()
    db.session.commit()
    flash(f'{year.label}已归档，所有考评数据已转为只读。', 'success')
    return redirect(url_for('main.scheme_settings_page'))


@main_bp.route('/admin/years/<int:year_id>/restore', methods=['POST'])
@admin_required
def restore_year(year_id):
    year = db.get_or_404(AcademicYear, year_id)
    if not _can_access_scheme(year.scheme_id):
        abort(403)
    AcademicYear.query.filter_by(scheme_id=year.scheme_id, status='active').update({'status': 'archived'})
    year.status = 'active'
    year.archived_at = None
    db.session.commit()
    flash(f'{year.label}已还原，可继续修改。', 'success')
    return redirect(url_for('main.scheme_settings_page'))


@main_bp.route('/admin/years/new', methods=['POST'])
@admin_required
def create_academic_year():
    if current_year():
        flash('请先归档当前学年，再开启新学年。', 'danger')
        return redirect(url_for('main.scheme_settings_page'))
    name = request.form.get('name', '').strip()
    source_id = request.form.get('source_year_id')
    if not name:
        flash('学年名称不能为空。', 'danger')
        return redirect(url_for('main.scheme_settings_page'))
    scheme = current_scheme()
    year = AcademicYear(name=_unique_year_name(name, scheme.code), display_name=name, scheme_id=scheme.id, status='active')
    db.session.add(year)
    db.session.flush()
    source = db.session.get(AcademicYear, source_id)
    if source:
        _clone_dictionary(source, year)
    db.session.commit()
    flash(f'{name}已开启，考评分数从零开始。', 'success')
    return redirect(url_for('main.scheme_settings_page'))


@main_bp.route('/admin/mobile-access')
@admin_required
def mobile_access_page():
    base_url = request.host_url.rstrip('/')
    modules = [
        {'key': 'admin-entry', 'name': '管理员录入', 'description': '选择教师和指标，录入客观数据或主观评分。', 'path': url_for('main.admin_entry_page'), 'tone': 'purple'},
        {'key': 'review', 'name': '一级审核', 'description': '在手机上查看教师材料，执行通过或退回。', 'path': url_for('main.review_page'), 'tone': 'yellow'},
        {'key': 'teacher-entry', 'name': '教师填报', 'description': '教师选择可自行录入的指标并提交材料。', 'path': url_for('main.teacher_entry_page'), 'tone': 'green'},
        {'key': 'teacher-results', 'name': '教师查看', 'description': '教师查看本人分项分、记录状态和总分。', 'path': url_for('main.my_results_page'), 'tone': 'blue'},
    ]
    return render_template('mobile_access.html', base_url=base_url, modules=modules)


@main_bp.route('/api/records', methods=['POST'])
@login_required
def create_record_api():
    year = current_year()
    if not year or year.status != 'active':
        return jsonify({'ok': False, 'message': '当前没有可录入的学年'}), 409
    payload = request.get_json(silent=True) or {}
    indicator = db.session.get(Indicator, payload.get('indicator_id'))
    if not indicator or indicator.academic_year_id != year.id or not indicator.enabled:
        return jsonify({'ok': False, 'message': '考评指标不存在或已经停用'}), 404

    is_admin = current_user.role in {'superadmin', 'admin'}
    if not is_admin and indicator.data_source != 'teacher':
        abort(403)
    if is_admin:
        target_user = db.session.get(User, payload.get('target_user_id'))
        if not target_user or target_user.role != 'teacher' or target_user not in _scheme_teachers(year.scheme):
            return jsonify({'ok': False, 'message': '请选择被考评教师'}), 400
    else:
        target_user = current_user

    if not indicator.allow_multiple_records and not indicator.requires_evidence:
        existing = EvaluationRecord.query.filter(
            EvaluationRecord.academic_year_id == year.id,
            EvaluationRecord.indicator_id == indicator.id,
            EvaluationRecord.target_user_id == target_user.id,
            EvaluationRecord.status != 'rejected',
        ).first()
        if existing:
            return jsonify({'ok': False, 'message': '该教师在此指标下已有记录，可在考评记录中调整分数。'}), 409

    inputs = payload.get('inputs') or {}
    if indicator.requires_evidence and indicator.scoring_type == 'count_score':
        inputs['count'] = 1
    missing = _missing_required_entry_field(indicator, inputs, payload.get('secondary_tracking_value'))
    if missing:
        return jsonify({'ok': False, 'message': f'请填写{missing}。'}), 400
    try:
        auto_score = calculate_score(indicator, inputs)
    except (TypeError, ValueError, KeyError) as error:
        return jsonify({'ok': False, 'message': f'无法计算分数：{error}'}), 400

    status = 'approved' if is_admin else 'pending'
    record = EvaluationRecord(
        scheme_id=year.scheme_id,
        academic_year_id=year.id,
        indicator_id=indicator.id,
        target_user_id=target_user.id,
        submitted_by_user_id=current_user.id,
        source='admin' if is_admin else 'teacher',
        status=status,
        input_json=json.dumps(inputs, ensure_ascii=False),
        secondary_tracking_value=str(payload.get('secondary_tracking_value') or '').strip() or None,
        auto_score=auto_score,
        final_score=auto_score if is_admin else None,
        note=str(payload.get('note') or '').strip(),
    )
    db.session.add(record)
    db.session.flush()
    db.session.add(AuditLog(
        user_id=current_user.id,
        action='record.create',
        entity_type='evaluation_record',
        entity_id=str(record.id),
        detail_json=json.dumps({'indicator_code': indicator.code, 'auto_score': auto_score}, ensure_ascii=False),
    ))
    db.session.commit()
    return jsonify({
        'ok': True,
        'record_id': record.id,
        'status': record.status,
        'auto_score': auto_score,
        'message': '已直接计入分数' if is_admin else '已提交，等待审核',
    })


@main_bp.route('/api/records/<int:record_id>', methods=['GET', 'PATCH'])
@login_required
def teacher_record_api(record_id):
    record = db.get_or_404(EvaluationRecord, record_id)
    if current_user.role != 'teacher' or record.source != 'teacher' or record.submitted_by_user_id != current_user.id:
        abort(403)
    if not _can_access_scheme(record.scheme_id):
        abort(403)
    if record.status not in {'pending', 'rejected'} or record.academic_year.status != 'active':
        return jsonify({'ok': False, 'message': '只有待审核或已退回的填报可以修改。'}), 409
    if request.method == 'GET':
        try:
            inputs = json.loads(record.input_json or '{}')
        except json.JSONDecodeError:
            inputs = {}
        return jsonify({
            'ok': True, 'record_id': record.id, 'indicator_id': record.indicator_id,
            'inputs': inputs, 'secondary_tracking_value': record.secondary_tracking_value or '',
            'note': record.note or '',
            'attachment': ({'id': record.attachments[0].id, 'name': record.attachments[0].original_name}
                           if record.attachments else None),
        })

    payload = request.get_json(silent=True) or {}
    if payload.get('indicator_id') != record.indicator_id:
        return jsonify({'ok': False, 'message': '编辑时不能更换考评指标。'}), 400
    inputs = payload.get('inputs') or {}
    if record.indicator.requires_evidence and record.indicator.scoring_type == 'count_score':
        inputs['count'] = 1
    missing = _missing_required_entry_field(record.indicator, inputs, payload.get('secondary_tracking_value'))
    if missing:
        return jsonify({'ok': False, 'message': f'请填写{missing}。'}), 400
    try:
        auto_score = calculate_score(record.indicator, inputs)
    except (TypeError, ValueError, KeyError) as error:
        return jsonify({'ok': False, 'message': f'无法计算分数：{error}'}), 400
    previous_status = record.status
    old_data = {
        'inputs': json.loads(record.input_json or '{}'),
        'secondary_tracking_value': record.secondary_tracking_value,
        'note': record.note,
        'auto_score': record.auto_score,
    }
    record.input_json = json.dumps(inputs, ensure_ascii=False)
    record.secondary_tracking_value = str(payload.get('secondary_tracking_value') or '').strip() or None
    record.note = str(payload.get('note') or '').strip()
    record.auto_score = auto_score
    record.status = 'pending'
    record.final_score = None
    record.review_note = ''
    record.reviewed_by_user_id = None
    record.reviewed_at = None
    db.session.add(AuditLog(
        user_id=current_user.id,
        action='record.resubmit' if previous_status == 'rejected' else 'record.pending_update',
        entity_type='evaluation_record',
        entity_id=str(record.id), detail_json=json.dumps({'before': old_data, 'auto_score': auto_score}, ensure_ascii=False),
    ))
    db.session.commit()
    message = '修改已重新提交，等待审核' if previous_status == 'rejected' else '修改已保存，仍等待审核'
    return jsonify({'ok': True, 'record_id': record.id, 'auto_score': auto_score, 'message': message})


def _missing_required_entry_field(indicator, inputs, tracking_value):
    try:
        rule = json.loads(indicator.scoring_rule_json or '{}')
    except json.JSONDecodeError:
        rule = {}
    for field in rule.get('extra_fields', []):
        if isinstance(field, dict) and field.get('required') and not str(inputs.get(field.get('key')) or '').strip():
            return field.get('label') or field.get('key') or '必填字段'
    try:
        tracking = json.loads(indicator.secondary_tracking_json or '{}')
    except json.JSONDecodeError:
        tracking = {}
    if tracking.get('enabled') and tracking.get('required') and not str(tracking_value or '').strip():
        return tracking.get('label') or '班级'
    return None


@main_bp.route('/api/records/<int:record_id>/attachments', methods=['POST'])
@login_required
def upload_record_attachments(record_id):
    from PIL import Image
    record = db.get_or_404(EvaluationRecord, record_id)
    if not _can_access_scheme(record.scheme_id) and record.submitted_by_user_id != current_user.id:
        abort(403)
    if current_user.role not in {'superadmin', 'admin'} and record.submitted_by_user_id != current_user.id:
        abort(403)
    if record.academic_year.status != 'active':
        return jsonify({'ok': False, 'message': '已归档学年不能上传材料。'}), 409
    if current_user.role == 'teacher' and record.status not in {'pending', 'rejected'}:
        return jsonify({'ok': False, 'message': '该申报当前状态不允许上传或替换材料。'}), 409
    files = request.files.getlist('files')
    if not files:
        return jsonify({'ok': False, 'message': '请选择图片。'}), 400
    if len(files) != 1:
        return jsonify({'ok': False, 'message': '每条评价记录只能上传 1 张材料图片，请将多份材料分条录入。'}), 400
    replace_existing = request.form.get('replace') == 'true'
    if record.attachments and not replace_existing:
        return jsonify({'ok': False, 'message': '这条记录已有材料图片；如有其他材料，请另建一条记录。'}), 409
    if replace_existing and current_user.role == 'teacher' and record.status != 'pending':
        return jsonify({'ok': False, 'message': '审核完成后不能替换材料图片。'}), 409
    old_attachments = list(record.attachments) if replace_existing else []
    folder = Path(current_app.config['UPLOAD_FOLDER']) / str(record.academic_year_id) / str(record.id)
    folder.mkdir(parents=True, exist_ok=True)
    uploaded = []
    for source in files[:1]:
        if not source.filename:
            continue
        try:
            image = _load_supported_image(source.stream)
            image.thumbnail((1800, 1800))
            if image.mode not in {'RGB', 'L'}:
                background = Image.new('RGB', image.size, 'white')
                if 'A' in image.getbands():
                    background.paste(image, mask=image.getchannel('A'))
                else:
                    background.paste(image)
                image = background
            elif image.mode == 'L':
                image = image.convert('RGB')
            stored_name = f'{uuid.uuid4().hex}.jpg'
            absolute = folder / stored_name
            image.save(absolute, 'JPEG', quality=82, optimize=True)
            relative = absolute.relative_to(Path(current_app.config['UPLOAD_FOLDER']))
            attachment = RecordAttachment(
                record_id=record.id, original_name=source.filename[:255], stored_path=str(relative),
                mime_type='image/jpeg', size_bytes=absolute.stat().st_size,
                width=image.width, height=image.height,
            )
            db.session.add(attachment)
            db.session.flush()
            uploaded.append({'id': attachment.id, 'name': attachment.original_name, 'size': attachment.size_bytes})
        except ValueError as error:
            return jsonify({'ok': False, 'message': str(error)}), 415
        except Exception:
            return jsonify({'ok': False, 'message': '图片读取或压缩失败，请更换图片后重试。'}), 400
    if not uploaded:
        return jsonify({'ok': False, 'message': '图片格式不受支持或文件已损坏。'}), 400
    for old in old_attachments:
        old_path = Path(current_app.config['UPLOAD_FOLDER']) / old.stored_path
        db.session.delete(old)
        try:
            old_path.unlink(missing_ok=True)
        except OSError:
            current_app.logger.warning('Unable to delete replaced attachment %s', old_path)
    db.session.commit()
    return jsonify({'ok': True, 'attachments': uploaded, 'message': f'已压缩上传 {len(uploaded)} 张图片'})


@main_bp.route('/attachments/<int:attachment_id>')
@login_required
def view_attachment(attachment_id):
    attachment = db.get_or_404(RecordAttachment, attachment_id)
    record = attachment.record
    if record.status == 'voided' and record.submitted_by_user_id != current_user.id:
        abort(404)
    if current_user.role in {'superadmin', 'admin', 'reviewer'} and not _can_access_scheme(record.scheme_id):
        abort(403)
    if current_user.role not in {'superadmin', 'admin', 'reviewer'} and record.target_user_id != current_user.id:
        abort(403)
    path = Path(attachment.stored_path)
    return send_from_directory(Path(current_app.config['UPLOAD_FOLDER']) / path.parent, path.name)


@main_bp.route('/api/attachments/<int:attachment_id>/recognize', methods=['POST'])
@login_required
def recognize_attachment(attachment_id):
    attachment = db.get_or_404(RecordAttachment, attachment_id)
    record = attachment.record
    if record.status == 'voided':
        if record.submitted_by_user_id != current_user.id:
            abort(404)
        return jsonify({'ok': False, 'message': '已作废记录不能再进行 AI 识别。'}), 409
    if current_user.role not in {'superadmin', 'admin'} and record.submitted_by_user_id != current_user.id:
        abort(403)
    if current_user.role == 'teacher' and record.status not in {'pending', 'rejected'}:
        return jsonify({'ok': False, 'message': '当前状态不能再修改识别内容。'}), 409
    setting = _ai_setting_for('vision')
    if not _ai_setting_ready(setting):
        return jsonify({'ok': False, 'message': '管理员尚未完成视觉识别模型配置。'}), 409
    absolute = Path(current_app.config['UPLOAD_FOLDER']) / attachment.stored_path
    try:
        encoded = base64.b64encode(absolute.read_bytes()).decode('ascii')
        prompt = (
            '请识别这张证书或获奖材料。用简洁中文输出：证书/奖项名称、获奖人、'
            '级别（校/区/市/省/国家）、奖次、颁发单位、日期。无法确认的字段写“未识别”，不要猜测。'
        )
        content = _call_openai(setting, [{'role': 'user', 'content': [
            {'type': 'text', 'text': prompt},
            {'type': 'image_url', 'image_url': {'url': f'data:image/jpeg;base64,{encoded}'}},
        ]}], max_tokens=500)
    except Exception:
        return jsonify({'ok': False, 'message': 'AI 识别失败，可直接手动填写。'}), 502
    record.note = ((record.note + '\n') if record.note else '') + f'[AI识别] {content}'
    db.session.commit()
    return jsonify({'ok': True, 'content': content})


def _record_input_fields(record, inputs):
    try:
        rule = json.loads(record.indicator.scoring_rule_json or '{}')
    except json.JSONDecodeError:
        rule = {}
    labels = {
        'qualified': '是否符合条件', 'option': '评价档次', 'level': '获奖级别',
        'rank': '奖次', 'score': '实际得分', 'count': '次数/材料数量',
        'value': '统计数值', 'years': '班主任任职年限',
    }
    for field in rule.get('extra_fields', []):
        if isinstance(field, dict) and field.get('key'):
            labels[field['key']] = field.get('label') or field['key']
    fields = []
    for key, raw_value in inputs.items():
        value = raw_value
        if key == 'qualified':
            value = '符合' if raw_value else '不符合'
        elif key == 'option':
            row = next((item for item in rule.get('options', []) if item.get('value') == raw_value), None)
            value = row.get('label', raw_value) if row else raw_value
        elif key in {'level', 'rank'}:
            label_key = f'{key}_label'
            row = next((item for item in rule.get('scores', []) if item.get(key) == raw_value), None)
            value = row.get(label_key, raw_value) if row else raw_value
        fields.append({'key': key, 'label': labels.get(key, key), 'value': value})
    return fields


def _record_view(record):
    try:
        inputs = json.loads(record.input_json or '{}')
    except json.JSONDecodeError:
        inputs = {}
    input_fields = _record_input_fields(record, inputs)
    input_text = '、'.join(f'{item["label"]}：{item["value"]}' for item in input_fields) or '无'
    try:
        tracking_config = json.loads(record.indicator.secondary_tracking_json or '{}')
    except json.JSONDecodeError:
        tracking_config = {}
    return {
        'id': record.id,
        'teacher': record.target_user.display_name,
        'scheme_name': record.scheme.name if record.scheme else '',
        'scheme_code': record.scheme.code if record.scheme else '',
        'academic_year': record.academic_year.label if record.academic_year else '',
        'indicator': record.indicator.name,
        'category': record.indicator.category.name,
        'source': record.source,
        'status': record.status,
        'status_label': {
            'draft': '草稿', 'pending': '待审核', 'approved': '已通过', 'rejected': '已退回',
            'voided': '已作废',
        }.get(record.status, record.status),
        'inputs': input_text,
        'input_fields': input_fields,
        'tracking': record.secondary_tracking_value,
        'tracking_label': tracking_config.get('label') or '跟踪字段',
        'auto_score': record.auto_score,
        'final_score': record.final_score,
        'overridden': record.admin_overridden,
        'note': record.note,
        'review_note': record.review_note,
        'created_at': record.created_at,
        'attachments': [{'id': item.id, 'name': item.original_name} for item in record.attachments],
        'can_edit': (
            record.source == 'teacher' and record.status in {'pending', 'rejected'}
            and record.academic_year.status == 'active'
        ),
        'can_void': (
            record.source == 'teacher' and record.status in {'pending', 'rejected'}
            and record.academic_year.status == 'active'
        ),
        'can_delete': (
            record.source == 'teacher' and record.status == 'voided'
            and record.academic_year.status == 'active'
        ),
    }


def _scheme_years():
    scheme = current_scheme()
    query = AcademicYear.query
    if scheme:
        query = query.filter_by(scheme_id=scheme.id)
    return query.order_by(AcademicYear.id.desc()).all()


def _selected_year():
    year_id = request.args.get('year_id', type=int)
    if year_id:
        year = db.session.get(AcademicYear, year_id)
        if year and _can_access_scheme(year.scheme_id):
            return year
    return current_year() or (_scheme_years()[0] if _scheme_years() else None)


def _clone_dictionary(source, target):
    for source_category in Category.query.filter_by(academic_year_id=source.id).order_by(Category.sort_order, Category.id):
        category = Category(
            academic_year_id=target.id, code=source_category.code, name=source_category.name,
            description=source_category.description, sort_order=source_category.sort_order,
            max_score=source_category.max_score,
        )
        db.session.add(category)
        db.session.flush()
        for source_item in source_category.indicators:
            db.session.add(Indicator(
                academic_year_id=target.id, category_id=category.id, code=source_item.code, name=source_item.name,
                description=source_item.description, sort_order=source_item.sort_order,
                score_group=source_item.score_group, data_source=source_item.data_source,
                scoring_type=source_item.scoring_type, scoring_rule_json=source_item.scoring_rule_json,
                allow_multiple_records=source_item.allow_multiple_records,
                requires_evidence=source_item.requires_evidence, ai_enabled=source_item.ai_enabled,
                secondary_tracking_json=source_item.secondary_tracking_json, enabled=source_item.enabled,
            ))


def _int_form(name, default=0):
    try:
        return int(request.form.get(name, default))
    except (TypeError, ValueError):
        return default


def _optional_float_form(name):
    value = request.form.get(name, '').strip()
    if not value:
        return None
    try:
        return float(value)
    except ValueError:
        return None


def _indicator_form_data(rule, tracking):
    return {
        'code': request.form.get('code', '').strip().lower(),
        'name': request.form.get('name', '').strip(),
        'description': request.form.get('description', '').strip(),
        'order': _int_form('sort_order', 0),
        'score_group': request.form.get('score_group', 'base'),
        'data_source': request.form.get('data_source', 'admin'),
        'scoring_type': request.form.get('scoring_type', 'manual_score'),
        'scoring_rule': rule,
        'allow_multiple_records': request.form.get('allow_multiple_records') == 'on',
        'requires_evidence': request.form.get('requires_evidence') == 'on',
        'ai_enabled': request.form.get('ai_enabled') == 'on',
        'secondary_tracking': tracking,
        'enabled': True,
    }


def _normalize_api_base(value):
    import re
    base = (value or '').strip().rstrip('/')
    if base and not base.startswith(('http://', 'https://')):
        base = 'http://' + base
    if base and not re.search(r'/v\d+(?:beta\d*)?$', base):
        base += '/v1'
    return base


def _uploaded_image_data_url(upload):
    from PIL import Image
    try:
        image = _load_supported_image(upload.stream)
        image.thumbnail((1800, 1800))
        if image.mode != 'RGB':
            if 'A' in image.getbands():
                background = Image.new('RGB', image.size, 'white')
                background.paste(image, mask=image.getchannel('A'))
                image = background
            else:
                image = image.convert('RGB')
        output = BytesIO()
        image.save(output, 'JPEG', quality=82, optimize=True)
    except ValueError:
        raise
    except Exception as error:
        raise ValueError('图片读取或压缩失败，请更换图片后重试。') from error
    encoded = base64.b64encode(output.getvalue()).decode('ascii')
    return f'data:image/jpeg;base64,{encoded}'


def _load_supported_image(stream):
    from PIL import Image, ImageOps
    try:
        from pillow_heif import register_heif_opener
        register_heif_opener()
    except ImportError:
        pass
    try:
        image = Image.open(stream)
        image_format = str(image.format or '').upper()
        if image_format not in {'JPEG', 'PNG', 'WEBP', 'HEIF', 'HEIC'}:
            raise ValueError('附件只能上传 JPG、PNG、WEBP、HEIC 或 HEIF 图片，不能上传 PDF 或其他文件。')
        image = ImageOps.exif_transpose(image)
        image.load()
        return image
    except ValueError:
        raise
    except Exception as error:
        raise ValueError('附件不是可识别的图片，不能上传 PDF 或其他文件。') from error


def _entry_recognition_prompt(indicator):
    rule = json.loads(indicator.scoring_rule_json or '{}')
    tracking = json.loads(indicator.secondary_tracking_json or '{}')
    field_instruction = ''
    if indicator.scoring_type in {'fixed_score', 'fixed_bonus'}:
        field_instruction = 'inputs.qualified：材料能证明符合该指标时为 true，明确不符合时为 false，无法判断时为 null。'
    elif indicator.scoring_type == 'tiered_score':
        choices = [{'value': row.get('value'), 'label': row.get('label')} for row in rule.get('options', [])]
        field_instruction = f'inputs.option：只能填写下列选项的 value，无法判断填 null：{json.dumps(choices, ensure_ascii=False)}。'
    elif indicator.scoring_type == 'matrix_score':
        choices = [{
            'level': row.get('level'), 'level_label': row.get('level_label', row.get('level')),
            'rank': row.get('rank'), 'rank_label': row.get('rank_label', row.get('rank')),
        } for row in rule.get('scores', [])]
        field_instruction = f'inputs.level 和 inputs.rank：只能使用下列合法组合中的值，无法判断填 null：{json.dumps(choices, ensure_ascii=False)}。'
    elif indicator.scoring_type == 'manual_score':
        field_instruction = f'inputs.score：仅当材料可直接确定分数时填写 {rule.get("min_score", 0)} 到 {rule.get("max_score", 0)} 之间的数值，否则填 null。'
    elif indicator.scoring_type == 'range_score':
        field_instruction = 'inputs.value：提取材料中用于本指标计分的数值，无法判断填 null。'
    elif indicator.scoring_type == 'count_deduction':
        field_instruction = 'inputs.count：提取材料明确证明的次数，必须是非负整数，无法判断填 null。'
    elif indicator.scoring_type == 'tenure_score':
        field_instruction = 'inputs.years：提取材料明确证明的任职年数，必须是非负整数，无法判断填 null。'
    elif indicator.scoring_type == 'count_score':
        field_instruction = 'inputs.count 固定填写 1，代表本次上传的一份材料。'
    else:
        field_instruction = 'inputs 使用空对象。'

    extra_instructions = []
    for field in rule.get('extra_fields', []):
        if not isinstance(field, dict) or not field.get('key'):
            continue
        options = field.get('options') or []
        option_text = f'，只能从这些选项中选择：{json.dumps(options, ensure_ascii=False)}' if options else ''
        extra_instructions.append(
            f'inputs.{field["key"]}：识别“{field.get("label") or field["key"]}”{option_text}，无法判断填 null。'
        )

    tracking_instruction = 'secondary_tracking_value 填 null。'
    if tracking.get('enabled'):
        label = tracking.get('label') or '补充字段'
        options = tracking.get('options') or []
        option_text = f'，若能识别只能从这些选项中选择：{json.dumps(options, ensure_ascii=False)}' if options else ''
        tracking_instruction = f'secondary_tracking_value：识别“{label}”{option_text}，无法判断填 null。'

    return (
        '你是学校考评材料识别助手。请根据图片和当前指标提取表单字段，不要猜测。'
        '只输出一个标准 JSON 对象，不要 Markdown、解释或额外文字。\n'
        f'指标名称：{indicator.name}\n指标说明：{indicator.description or "无"}\n'
        f'计分类型：{indicator.scoring_type}\n{field_instruction}\n' + '\n'.join(extra_instructions) + f'\n{tracking_instruction}\n'
        'note：用简洁中文汇总材料中能确认的信息，包括材料/奖项名称、相关人员、级别、奖次、颁发单位和日期；'
        '无法识别的关键字段写“未识别”，不要虚构。\n'
        '严格返回：{"inputs": {...}, "secondary_tracking_value": null, "note": "..."}'
    )


def _sanitize_entry_recognition(indicator, recognized):
    if not isinstance(recognized, dict):
        raise ValueError('AI 识别结果不是有效对象。')
    raw_inputs = recognized.get('inputs') if isinstance(recognized.get('inputs'), dict) else {}
    rule = json.loads(indicator.scoring_rule_json or '{}')
    tracking = json.loads(indicator.secondary_tracking_json or '{}')
    mapped = {}
    filled = []

    def choice_value(candidate, rows, value_key='value', label_key='label'):
        text = str(candidate or '').strip().lower()
        if not text:
            return None
        for row in rows:
            value = str(row.get(value_key) or '').strip()
            label = str(row.get(label_key) or '').strip()
            if text in {value.lower(), label.lower()}:
                return value
        return None

    scoring_type = indicator.scoring_type
    if scoring_type in {'fixed_score', 'fixed_bonus'}:
        value = raw_inputs.get('qualified')
        if isinstance(value, bool):
            mapped['qualified'] = value
            filled.append('是否符合条件')
        elif str(value).strip().lower() in {'true', '是', '符合', '1'}:
            mapped['qualified'] = True
            filled.append('是否符合条件')
        elif str(value).strip().lower() in {'false', '否', '不符合', '0'}:
            mapped['qualified'] = False
            filled.append('是否符合条件')
    elif scoring_type == 'tiered_score':
        value = choice_value(raw_inputs.get('option'), rule.get('options', []))
        if value is not None:
            mapped['option'] = value
            filled.append('评价档次')
    elif scoring_type == 'matrix_score':
        rows = rule.get('scores', [])
        level = choice_value(raw_inputs.get('level'), rows, 'level', 'level_label')
        rank = choice_value(raw_inputs.get('rank'), rows, 'rank', 'rank_label')
        if level is not None and rank is not None and any(row.get('level') == level and row.get('rank') == rank for row in rows):
            mapped.update({'level': level, 'rank': rank})
            filled.extend(['获奖级别', '奖次'])
    elif scoring_type in {'manual_score', 'range_score', 'count_deduction', 'tenure_score'}:
        key = {'manual_score': 'score', 'range_score': 'value', 'count_deduction': 'count', 'tenure_score': 'years'}[scoring_type]
        try:
            number = float(raw_inputs.get(key))
            if scoring_type in {'count_deduction', 'tenure_score'}:
                if number < 0 or not number.is_integer():
                    raise ValueError
                number = int(number)
            elif scoring_type == 'manual_score':
                if number < float(rule.get('min_score', number)) or number > float(rule.get('max_score', number)):
                    raise ValueError
            mapped[key] = number
            filled.append({'score': '实际得分', 'value': '统计数值', 'count': '次数', 'years': '任职年限'}[key])
        except (TypeError, ValueError):
            pass
    elif scoring_type == 'count_score':
        mapped['count'] = 1

    for field in rule.get('extra_fields', []):
        if not isinstance(field, dict) or not field.get('key'):
            continue
        key = field['key']
        candidate = str(raw_inputs.get(key) or '').strip()
        options = field.get('options') or []
        if candidate and (not options or candidate in options):
            mapped[key] = candidate[:200]
            filled.append(field.get('label') or key)

    tracking_value = None
    if tracking.get('enabled'):
        candidate = str(recognized.get('secondary_tracking_value') or '').strip()
        options = tracking.get('options') or []
        if candidate and (not options or candidate in options):
            tracking_value = candidate[:200]
            filled.append(tracking.get('label') or '补充字段')
    note = str(recognized.get('note') or '').strip()[:2000]
    if not note:
        note = 'AI 已完成材料识别，请结合原图人工核对。'
    filled.append('补充说明')
    return {
        'inputs': mapped,
        'secondary_tracking_value': tracking_value,
        'note': note,
        'filled_fields': filled,
    }


def _call_openai(setting, messages, max_tokens=300):
    import requests
    headers = {'Content-Type': 'application/json'}
    if setting.api_key:
        headers['Authorization'] = f'Bearer {setting.api_key}'
    response = requests.post(
        _normalize_api_base(setting.api_base) + '/chat/completions',
        headers=headers,
        json={'model': setting.model_name, 'messages': messages, 'max_tokens': max_tokens, 'temperature': 0.1},
        timeout=120,
    )
    response.raise_for_status()
    return str(response.json()['choices'][0]['message']['content']).strip()


def _parse_json_object(content):
    text_content = (content or '').strip()
    if text_content.startswith('```'):
        text_content = text_content.split('\n', 1)[-1].rsplit('```', 1)[0].strip()
    start, end = text_content.find('{'), text_content.rfind('}')
    if start < 0 or end <= start:
        raise ValueError('AI 没有返回可解析的 JSON 对象。')
    try:
        return json.loads(text_content[start:end + 1])
    except json.JSONDecodeError as error:
        raise ValueError(f'AI 返回的 JSON 格式不正确：第 {error.lineno} 行。') from error


def _unique_year_name(display_name, scheme_code):
    candidate = display_name
    if not AcademicYear.query.filter_by(name=candidate).first():
        return candidate
    base = f'{display_name} @ {scheme_code}'
    candidate = base
    index = 2
    while AcademicYear.query.filter_by(name=candidate).first():
        candidate = f'{base}-{index}'
        index += 1
    return candidate


def _quote(value):
    from urllib.parse import quote

    return quote(value)
