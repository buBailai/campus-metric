from io import BytesIO
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


HEADERS = [
    '姓名', '登录账号', '工号', '角色', '方案编号', '初始密码',
    '班主任', '年段长', '班主任年限', '启用状态',
]
REQUIRED_HEADERS = {'姓名', '登录账号', '角色', '方案编号', '初始密码'}
ROLE_MAP = {
    '教师': 'teacher', 'teacher': 'teacher',
    '审核人': 'reviewer', '审核员': 'reviewer', 'reviewer': 'reviewer',
    '管理员': 'admin', '部门管理员': 'admin', 'admin': 'admin',
}


def _text(value):
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _boolean(value, default, row_number, field_name, errors):
    text = _text(value).lower()
    if not text:
        return default
    if text in {'是', '有', '启用', 'true', 'yes', 'y', '1'}:
        return True
    if text in {'否', '无', '停用', 'false', 'no', 'n', '0'}:
        return False
    errors.append(f'第{row_number}行“{field_name}”只能填写是/否')
    return default


def build_account_template(scheme_codes=None):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = '账号导入'
    sheet.append(HEADERS)
    example_scheme = '、'.join(scheme_codes[:2]) if scheme_codes else 'SCHEME-001'
    sheet.append(['张老师', 'zhanglaoshi', 'T001', '教师', example_scheme, '12345678', '是', '否', 3, '启用'])

    header_fill = PatternFill('solid', fgColor='8D68F6')
    for cell in sheet[1]:
        cell.font = Font(color='FFFFFF', bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    widths = [14, 18, 14, 14, 22, 18, 11, 11, 14, 12]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    sheet.freeze_panes = 'A2'
    sheet.auto_filter.ref = f'A1:J2'

    role_validation = DataValidation(type='list', formula1='"教师,审核人,部门管理员"', allow_blank=False)
    yes_no_validation = DataValidation(type='list', formula1='"是,否"', allow_blank=True)
    active_validation = DataValidation(type='list', formula1='"启用,停用"', allow_blank=True)
    sheet.add_data_validation(role_validation)
    sheet.add_data_validation(yes_no_validation)
    sheet.add_data_validation(active_validation)
    role_validation.add('D2:D501')
    yes_no_validation.add('G2:H501')
    active_validation.add('J2:J501')

    guide = workbook.create_sheet('填写说明')
    guide.append(['字段', '是否必填', '填写说明'])
    guide_rows = [
        ('姓名', '是', '系统显示姓名'),
        ('登录账号', '是', '全系统唯一，建议使用工号或拼音；导入时自动转为小写'),
        ('工号', '否', '填写后必须全系统唯一；含前导零时请按文本填写'),
        ('角色', '是', '教师、审核人或部门管理员'),
        ('方案编号', '是', '可填写一套或多套；多套方案用逗号、顿号或分号分隔，编号必须与“方案与学年”页面完全一致'),
        ('初始密码', '是', '至少 8 个字符'),
        ('班主任', '否', '是/否；仅教师有效，留空默认为是'),
        ('年段长', '否', '是/否；仅教师有效，留空默认为否'),
        ('班主任年限', '否', '非负整数；仅教师有效，留空默认为 0'),
        ('启用状态', '否', '启用/停用，留空默认为启用'),
    ]
    for row in guide_rows:
        guide.append(row)
    if scheme_codes:
        guide.append([])
        guide.append(['当前可用方案编号', '', '、'.join(scheme_codes)])
    for cell in guide[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='EEE9FF')
    guide.column_dimensions['A'].width = 18
    guide.column_dimensions['B'].width = 12
    guide.column_dimensions['C'].width = 78
    guide.freeze_panes = 'A2'

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def parse_account_workbook(upload, schemes, existing_usernames, existing_employee_numbers, max_rows=500):
    errors = []
    try:
        workbook = load_workbook(upload, read_only=True, data_only=True)
    except Exception:
        return [], ['文件无法读取，请使用系统提供的 .xlsx 模板']
    sheet = workbook['账号导入'] if '账号导入' in workbook.sheetnames else workbook.active
    rows = sheet.iter_rows(values_only=True)
    first_row = next(rows, None)
    if not first_row:
        return [], ['导入文件没有表头']

    headers = [_text(value) for value in first_row]
    missing = sorted(REQUIRED_HEADERS - set(headers))
    if missing:
        return [], [f'缺少必填列：{"、".join(missing)}']
    positions = {name: headers.index(name) for name in HEADERS if name in headers}

    scheme_by_code = {scheme.code: scheme for scheme in schemes}
    known_usernames = {value.lower() for value in existing_usernames}
    known_employee_numbers = {value for value in existing_employee_numbers if value}
    file_usernames = set()
    file_employee_numbers = set()
    records = []

    for row_number, values in enumerate(rows, start=2):
        def value(name):
            index = positions.get(name)
            return values[index] if index is not None and index < len(values) else None

        if not any(_text(item) for item in values):
            continue
        if len(records) >= max_rows:
            errors.append(f'一次最多导入 {max_rows} 个账号')
            break

        display_name = _text(value('姓名'))
        username = _text(value('登录账号')).lower()
        employee_no = _text(value('工号')) or None
        role_text = _text(value('角色')).lower()
        role = ROLE_MAP.get(role_text)
        scheme_code_text = _text(value('方案编号'))
        scheme_codes = []
        for code in re.split(r'[,，、;；\s]+', scheme_code_text):
            if code and code not in scheme_codes:
                scheme_codes.append(code)
        password = _text(value('初始密码'))
        tenure_text = _text(value('班主任年限'))

        if not display_name:
            errors.append(f'第{row_number}行缺少姓名')
        if not username:
            errors.append(f'第{row_number}行缺少登录账号')
        elif username in known_usernames or username in file_usernames:
            errors.append(f'第{row_number}行登录账号“{username}”重复')
        if employee_no and (employee_no in known_employee_numbers or employee_no in file_employee_numbers):
            errors.append(f'第{row_number}行工号“{employee_no}”重复')
        if not role:
            errors.append(f'第{row_number}行角色“{_text(value("角色"))}”无效')
        missing_scheme_codes = [code for code in scheme_codes if code not in scheme_by_code]
        if not scheme_codes:
            errors.append(f'第{row_number}行缺少方案编号')
        elif missing_scheme_codes:
            errors.append(f'第{row_number}行方案编号“{"、".join(missing_scheme_codes)}”不存在')
        if len(password) < 8:
            errors.append(f'第{row_number}行初始密码至少需要 8 个字符')
        try:
            tenure_years = int(float(tenure_text)) if tenure_text else 0
            if tenure_years < 0 or (tenure_text and float(tenure_text) != tenure_years):
                raise ValueError
        except ValueError:
            tenure_years = 0
            errors.append(f'第{row_number}行班主任年限必须是非负整数')

        is_homeroom_teacher = _boolean(value('班主任'), True, row_number, '班主任', errors)
        is_grade_leader = _boolean(value('年段长'), False, row_number, '年段长', errors)
        is_active = _boolean(value('启用状态'), True, row_number, '启用状态', errors)
        if role and role != 'teacher':
            is_homeroom_teacher = False
            is_grade_leader = False
            tenure_years = 0

        if username:
            file_usernames.add(username)
        if employee_no:
            file_employee_numbers.add(employee_no)
        records.append({
            'display_name': display_name, 'username': username, 'employee_no': employee_no,
            'role': role, 'scheme_code': scheme_codes[0] if scheme_codes else '',
            'scheme_codes': scheme_codes, 'password': password,
            'is_homeroom_teacher': is_homeroom_teacher, 'is_grade_leader': is_grade_leader,
            'tenure_years': tenure_years, 'is_active_flag': is_active,
        })

    if not records and not errors:
        errors.append('导入文件中没有账号数据')
    return records, errors
