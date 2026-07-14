import json
import unittest
import tempfile
from io import BytesIO
from pathlib import Path
from unittest.mock import patch

from PIL import Image
from openpyxl import Workbook, load_workbook
from werkzeug.security import generate_password_hash

from evaluation_app import create_app, db
from evaluation_app.dictionary_service import (
    blank_template, calculate_score, full_example, import_dictionary,
    validate_dictionary,
)
from evaluation_app.models import (
    AIModelSetting, AcademicYear, AuditLog, Category, EvaluationRecord, EvaluationScheme, Indicator,
    SchemeMembership, User,
)
from evaluation_app.update_service import apply_staged_update


class DictionaryTests(unittest.TestCase):
    def setUp(self):
        self.app = create_app({
            'TESTING': True,
            'SECRET_KEY': 'test-secret',
            'SQLALCHEMY_DATABASE_URI': 'sqlite:///:memory:',
        })
        self.context = self.app.app_context()
        self.context.push()
        db.drop_all()
        db.create_all()
        self.admin = User(
            username='admin', display_name='管理员', role='admin',
            password_hash=generate_password_hash('password123'),
        )
        self.year = AcademicYear(name='2026—2027学年', status='active')
        db.session.add_all([self.admin, self.year])
        db.session.commit()

    def tearDown(self):
        db.session.remove()
        db.drop_all()
        self.context.pop()

    def test_templates_are_valid(self):
        self.assertTrue(validate_dictionary(blank_template())['valid'])
        result = validate_dictionary(full_example())
        self.assertTrue(result['valid'], result['errors'])
        self.assertEqual(result['summary']['items'], 9)

    def test_duplicate_indicator_code_is_rejected(self):
        data = full_example()
        data['categories'][0]['items'][1]['code'] = data['categories'][0]['items'][0]['code']
        result = validate_dictionary(data)
        self.assertFalse(result['valid'])
        self.assertTrue(any('重复' in item['message'] for item in result['errors']))

    def test_update_rolls_back_database_before_code_replacement(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            (root / 'instance').mkdir()
            (root / 'instance' / 'evaluation.sqlite').write_bytes(b'original-database')
            (root / 'app.py').write_text('old', encoding='utf-8')
            staging = root / 'backups' / 'update_staging' / 'app_new'
            staging.mkdir(parents=True)
            (staging / 'app.py').write_text('new', encoding='utf-8')
            replaced, rollback = apply_staged_update(root)
            self.assertEqual(replaced, 1)
            self.assertEqual((rollback / 'instance' / 'evaluation.sqlite').read_bytes(), b'original-database')
            self.assertEqual((root / 'app.py').read_text(encoding='utf-8'), 'new')

    def test_import_and_score(self):
        result = import_dictionary(full_example(), self.admin.id)
        self.assertTrue(result['valid'])
        repeated = import_dictionary(full_example(), self.admin.id)
        self.assertTrue(repeated['valid'])
        indicator = Indicator.query.filter_by(code='count_minus').one()
        self.assertEqual(calculate_score(indicator, {'count': 2}), 3)
        matrix = Indicator.query.filter_by(code='matrix').one()
        self.assertEqual(calculate_score(matrix, {'level': 'district', 'rank': 'second'}), 2)


class RouteTests(unittest.TestCase):
    def setUp(self):
        self.tempdir = tempfile.TemporaryDirectory()
        self.app = create_app({
            'TESTING': True,
            'SECRET_KEY': 'test-secret',
            'SQLALCHEMY_DATABASE_URI': 'sqlite:///:memory:',
            'UPLOAD_FOLDER': self.tempdir.name,
        })
        self.context = self.app.app_context()
        self.context.push()
        db.drop_all()
        db.create_all()
        self.client = self.app.test_client()

    def tearDown(self):
        db.session.remove()
        db.drop_all()
        self.context.pop()
        self.tempdir.cleanup()

    def test_bulk_account_import_template_validation_and_atomic_write(self):
        self.client.post('/setup', data={
            'username': 'admin', 'display_name': '管理员',
            'password': 'password123', 'academic_year': '2026—2027学年',
        })
        scheme = EvaluationScheme.query.one()

        template_response = self.client.get('/admin/users/import-template')
        self.assertEqual(template_response.status_code, 200)
        template = load_workbook(BytesIO(template_response.data), read_only=True)
        self.assertIn('账号导入', template.sheetnames)
        self.assertIn('填写说明', template.sheetnames)
        self.assertEqual(template['账号导入']['A1'].value, '姓名')

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = '账号导入'
        sheet.append(['姓名', '登录账号', '工号', '角色', '方案编号', '初始密码', '班主任', '年段长', '班主任年限', '启用状态'])
        sheet.append(['李老师', 'TEACHER02', 'T002', '教师', scheme.code, 'password123', '是', '是', 5, '启用'])
        sheet.append(['审核老师', 'reviewer02', 'R002', '审核人', scheme.code, 'password123', '', '', '', '停用'])
        upload = BytesIO()
        workbook.save(upload)
        upload.seek(0)
        response = self.client.post(
            '/admin/users/import', data={'file': (upload, 'accounts.xlsx')},
            content_type='multipart/form-data', follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn('成功导入 2 个账号'.encode(), response.data)
        teacher = User.query.filter_by(username='teacher02').one()
        reviewer = User.query.filter_by(username='reviewer02').one()
        self.assertTrue(teacher.is_homeroom_teacher)
        self.assertTrue(teacher.is_grade_leader)
        self.assertEqual(teacher.tenure_years, 5)
        self.assertFalse(reviewer.is_active_flag)
        self.assertEqual(SchemeMembership.query.filter_by(user_id=teacher.id).one().membership_role, 'participant')
        self.assertEqual(SchemeMembership.query.filter_by(user_id=reviewer.id).one().membership_role, 'reviewer')
        self.assertEqual(AuditLog.query.filter_by(action='user.bulk_import').count(), 1)

        duplicate_workbook = Workbook()
        duplicate_sheet = duplicate_workbook.active
        duplicate_sheet.append(['姓名', '登录账号', '角色', '方案编号', '初始密码'])
        duplicate_sheet.append(['重复教师', 'teacher02', '教师', scheme.code, 'password123'])
        duplicate_sheet.append(['不应写入', 'teacher03', '教师', scheme.code, 'password123'])
        duplicate_upload = BytesIO()
        duplicate_workbook.save(duplicate_upload)
        duplicate_upload.seek(0)
        response = self.client.post(
            '/admin/users/import', data={'file': (duplicate_upload, 'duplicate.xlsx')},
            content_type='multipart/form-data', follow_redirects=True,
        )
        self.assertIn('导入未写入任何账号'.encode(), response.data)
        self.assertIsNone(User.query.filter_by(username='teacher03').first())

    def test_setup_import_and_dynamic_sources(self):
        response = self.client.post('/setup', data={
            'username': 'admin', 'display_name': '管理员',
            'password': 'password123', 'academic_year': '2026—2027学年',
        })
        self.assertEqual(response.status_code, 302)
        self.assertEqual(User.query.filter_by(username='admin').one().role, 'superadmin')

        response = self.client.post('/api/admin/dictionary/import', json={
            'confirmed': True,
            'dictionary': full_example(),
        })
        self.assertEqual(response.status_code, 200, response.get_json())

        mobile_page = self.client.get('/admin/mobile-access')
        self.assertEqual(mobile_page.status_code, 200)
        self.assertIn('移动访问'.encode(), mobile_page.data)
        self.assertIn('管理员录入'.encode(), mobile_page.data)
        self.assertIn('教师填报'.encode(), mobile_page.data)

        dictionary_page = self.client.get('/admin/dictionary')
        self.assertEqual(dictionary_page.status_code, 200)
        self.assertIn('mobile-bottom-nav'.encode(), dictionary_page.data)
        self.assertIn('具体计算方式'.encode(), dictionary_page.data)
        self.assertIn('编辑此指标'.encode(), dictionary_page.data)
        self.assertIn(b'indicator-detail-card', dictionary_page.data)
        first_indicator_id = Indicator.query.order_by(Indicator.id).first().id
        self.assertIn(f'indicator_id={first_indicator_id}'.encode(), dictionary_page.data)
        targeted_editor = self.client.get('/admin/dictionary/manual', query_string={'indicator_id': first_indicator_id})
        self.assertIn(f'id="indicator-{first_indicator_id}" open'.encode(), targeted_editor.data)

        for path, marker in [
            ('/admin/dictionary/manual', '手动配置'),
            ('/admin/settings/scheme', '方案与学年'),
            ('/admin/settings/ai', 'AI 大模型配置'),
            ('/admin/settings/update', '在线更新'),
            ('/admin/ranking', '总分排名'),
            ('/admin/archive-analytics', '档案跟踪'),
            ('/profile', '退出当前账号'),
        ]:
            page = self.client.get(path)
            self.assertEqual(page.status_code, 200, path)
            self.assertIn(marker.encode(), page.data)

        prefixed_update_page = self.client.get(
            '/admin/settings/update',
            environ_overrides={'SCRIPT_NAME': '/campus-evaluation/app'},
        )
        self.assertEqual(prefixed_update_page.status_code, 200)
        self.assertIn(b'window.APP_ROOT = "/campus-evaluation/app"', prefixed_update_page.data)
        self.assertIn(b'/campus-evaluation/app/api/admin/update/check', prefixed_update_page.data)

        admin_items = self.client.get('/api/entry-indicators?mode=admin').get_json()
        admin_codes = [item['code'] for category in admin_items['categories'] for item in category['indicators']]
        self.assertIn('manual', admin_codes)
        self.assertNotIn('fixed', admin_codes)

        teacher = User(
            username='teacher1', display_name='张老师', role='teacher',
            password_hash=generate_password_hash('password123'),
            scheme_id=EvaluationScheme.query.first().id,
        )
        db.session.add(teacher)
        db.session.commit()
        self.client.post('/logout')
        self.client.post('/login', data={'username': 'teacher1', 'password': 'password123'})

        teacher_entry_page = self.client.get('/entry/new')
        self.assertEqual(teacher_entry_page.status_code, 200)
        self.assertIn('AI 识别并自动填写'.encode(), teacher_entry_page.data)

        teacher_items = self.client.get('/api/entry-indicators').get_json()
        teacher_codes = [item['code'] for category in teacher_items['categories'] for item in category['indicators']]
        self.assertIn('fixed', teacher_codes)
        self.assertIn('matrix', teacher_codes)
        self.assertNotIn('manual', teacher_codes)

        fixed = Indicator.query.filter_by(code='fixed').one()
        ai_setting = AIModelSetting(
            provider='custom', api_base='https://ai.example.test/v1', api_key='test-key',
            model_name='vision-test', enabled=True,
        )
        db.session.add(ai_setting)
        db.session.commit()
        recognize_image = BytesIO()
        Image.new('RGB', (900, 600), '#eee9ff').save(recognize_image, 'JPEG')
        recognize_image.seek(0)
        with patch('evaluation_app.routes._call_openai', return_value='''{
          "inputs": {"qualified": true},
          "secondary_tracking_value": null,
          "note": "班主任培训合格证书，获证人张老师，颁发日期2026年6月。"
        }'''):
            recognized = self.client.post(
                '/api/entry/recognize',
                data={'indicator_id': str(fixed.id), 'file': (recognize_image, '培训证书.jpg')},
                content_type='multipart/form-data',
            )
        self.assertEqual(recognized.status_code, 200, recognized.get_json())
        self.assertTrue(recognized.get_json()['inputs']['qualified'])
        self.assertIn('培训合格证书', recognized.get_json()['note'])

        matrix_indicator = Indicator.query.filter_by(code='matrix').one()
        matrix_image = BytesIO()
        Image.new('RGB', (900, 600), '#e7f8ec').save(matrix_image, 'JPEG')
        matrix_image.seek(0)
        with patch('evaluation_app.routes._call_openai', return_value='''{
          "inputs": {"level": "区级", "rank": "二等奖"},
          "secondary_tracking_value": null,
          "note": "区级二等奖获奖证书，获奖人张老师。"
        }'''):
            recognized_matrix = self.client.post(
                '/api/entry/recognize',
                data={'indicator_id': str(matrix_indicator.id), 'file': (matrix_image, '获奖证书.jpg')},
                content_type='multipart/form-data',
            )
        self.assertEqual(recognized_matrix.status_code, 200, recognized_matrix.get_json())
        self.assertEqual(recognized_matrix.get_json()['inputs'], {'level': 'district', 'rank': 'second'})
        db.session.delete(ai_setting)
        db.session.commit()

        response = self.client.post('/api/records', json={
            'indicator_id': fixed.id,
            'inputs': {'qualified': True},
            'note': '培训证书',
        })
        payload = response.get_json()
        self.assertEqual(response.status_code, 200, payload)
        self.assertEqual(payload['status'], 'pending')
        self.assertEqual(payload['auto_score'], 2)

        image_data = BytesIO()
        Image.new('RGB', (2200, 1200), '#8d68f6').save(image_data, 'PNG')
        image_data.seek(0)
        response = self.client.post(
            f'/api/records/{payload["record_id"]}/attachments',
            data={'files': (image_data, '培训证书.png')},
            content_type='multipart/form-data',
        )
        attachment_payload = response.get_json()
        self.assertEqual(response.status_code, 200, attachment_payload)
        self.assertEqual(len(attachment_payload['attachments']), 1)
        self.assertLess(attachment_payload['attachments'][0]['size'], 200000)
        image_data = BytesIO()
        Image.new('RGB', (200, 200), '#4ccc75').save(image_data, 'PNG')
        image_data.seek(0)
        response = self.client.post(
            f'/api/records/{payload["record_id"]}/attachments',
            data={'files': (image_data, '第二份材料.png')}, content_type='multipart/form-data',
        )
        self.assertEqual(response.status_code, 409)

        count_add = Indicator.query.filter_by(code='count_add').one()
        pdf_record = self.client.post('/api/records', json={
            'indicator_id': count_add.id, 'inputs': {'count': 1}, 'note': '非法附件测试',
        }).get_json()
        pdf_response = self.client.post(
            f'/api/records/{pdf_record["record_id"]}/attachments',
            data={'files': (BytesIO(b'%PDF-1.4 fake pdf'), '伪装材料.jpg')},
            content_type='multipart/form-data',
        )
        self.assertEqual(pdf_response.status_code, 415)
        self.assertIn('不能上传 PDF', pdf_response.get_json()['message'])

        matrix_rule = json.loads(matrix_indicator.scoring_rule_json)
        matrix_rule['extra_fields'] = [{'key': 'student_name', 'label': '学生姓名', 'input_type': 'text', 'required': True}]
        matrix_indicator.scoring_rule_json = json.dumps(matrix_rule, ensure_ascii=False)
        matrix_indicator.secondary_tracking_json = '{"enabled": true, "label": "班级", "required": true, "input_type": "text"}'
        db.session.commit()
        matrix_record = self.client.post('/api/records', json={
            'indicator_id': matrix_indicator.id,
            'inputs': {'level': 'district', 'rank': 'second', 'student_name': '学生甲'},
            'secondary_tracking_value': '三年1班', 'note': '区级二等奖材料',
        })
        self.assertEqual(matrix_record.status_code, 200, matrix_record.get_json())
        resubmit_record = self.client.post('/api/records', json={
            'indicator_id': count_add.id, 'inputs': {'count': 1}, 'note': '等待退回测试',
        })
        self.assertEqual(resubmit_record.status_code, 200, resubmit_record.get_json())

        result_page = self.client.get('/my/results')
        self.assertEqual(result_page.status_code, 200)
        self.assertIn(b'CampusMetric', result_page.data)
        self.assertIn('校园评价系统'.encode(), result_page.data)
        self.assertIn('待审核'.encode(), result_page.data)
        self.assertIn('编辑待审核填报'.encode(), result_page.data)

        edit_page = self.client.get(f'/entry/{payload["record_id"]}/edit')
        self.assertEqual(edit_page.status_code, 200)
        self.assertIn('修改待审核填报'.encode(), edit_page.data)
        edit_response = self.client.patch(f'/api/records/{payload["record_id"]}', json={
            'indicator_id': fixed.id, 'inputs': {'qualified': True},
            'secondary_tracking_value': '', 'note': '教师修改后的培训证书说明',
        })
        self.assertEqual(edit_response.status_code, 200, edit_response.get_json())
        self.assertEqual(db.session.get(EvaluationRecord, payload['record_id']).note, '教师修改后的培训证书说明')

        response = self.client.post('/profile/password', data={
            'current_password': 'wrong-password',
            'new_password': 'newpassword456',
            'confirm_password': 'newpassword456',
        }, follow_redirects=True)
        self.assertIn('当前密码不正确'.encode(), response.data)
        response = self.client.post('/profile/password', data={
            'current_password': 'password123',
            'new_password': 'newpassword456',
            'confirm_password': 'newpassword456',
        }, follow_redirects=True)
        self.assertIn('密码修改成功'.encode(), response.data)

        self.client.post('/logout')
        response = self.client.post('/login', data={
            'username': 'teacher1', 'password': 'newpassword456',
        })
        self.assertEqual(response.status_code, 302)
        self.client.post('/logout')
        self.client.post('/login', data={'username': 'admin', 'password': 'password123'})
        response = self.client.post('/admin/users', data={
            'display_name': '审核老师', 'username': 'reviewer1',
            'password': 'password123', 'role': 'reviewer',
        }, follow_redirects=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn('审核老师'.encode(), response.data)

        self.client.post('/logout')
        self.client.post('/login', data={'username': 'reviewer1', 'password': 'password123'})
        review_page = self.client.get('/review')
        self.assertEqual(review_page.status_code, 200)
        self.assertIn('张老师'.encode(), review_page.data)
        self.assertIn('获奖级别'.encode(), review_page.data)
        self.assertIn('区级'.encode(), review_page.data)
        self.assertIn('奖次'.encode(), review_page.data)
        self.assertIn('二等奖'.encode(), review_page.data)
        self.assertIn('学生姓名'.encode(), review_page.data)
        self.assertIn('学生甲'.encode(), review_page.data)
        self.assertIn('班级'.encode(), review_page.data)
        self.assertIn('三年1班'.encode(), review_page.data)
        self.assertIn(b'<details class="record-card review-record-card">', review_page.data)
        self.assertIn('展开完整内容'.encode(), review_page.data)
        self.assertIn(b'data-attachment-preview', review_page.data)
        self.assertNotIn(b'target="_blank"', review_page.data)
        response = self.client.post(
            f'/review/{payload["record_id"]}',
            data={'action': 'approve', 'review_note': '材料无误'},
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        record = db.session.get(EvaluationRecord, payload['record_id'])
        self.assertEqual(record.status, 'approved')
        self.assertEqual(record.final_score, 2)
        matrix_record_id = matrix_record.get_json()['record_id']
        matrix_approved = self.client.post(
            f'/review/{matrix_record_id}', data={'action': 'approve', 'review_note': '学生获奖材料无误'},
            follow_redirects=True,
        )
        self.assertEqual(matrix_approved.status_code, 200)
        rejected_id = resubmit_record.get_json()['record_id']
        rejected = self.client.post(
            f'/review/{rejected_id}', data={'action': 'reject', 'review_note': '请补充材料说明'},
            follow_redirects=True,
        )
        self.assertEqual(rejected.status_code, 200)

        self.client.post('/logout')
        self.client.post('/login', data={'username': 'teacher1', 'password': 'newpassword456'})
        rejected_results = self.client.get('/my/results')
        self.assertIn('修改并重新提交'.encode(), rejected_results.data)
        resubmitted = self.client.patch(f'/api/records/{rejected_id}', json={
            'indicator_id': count_add.id, 'inputs': {'count': 1},
            'secondary_tracking_value': '', 'note': '已补充材料说明并重新提交',
        })
        self.assertEqual(resubmitted.status_code, 200, resubmitted.get_json())
        self.assertEqual(db.session.get(EvaluationRecord, rejected_id).status, 'pending')
        self.assertEqual(AuditLog.query.filter_by(action='record.resubmit', entity_id=str(rejected_id)).count(), 1)
        locked_edit = self.client.patch(f'/api/records/{payload["record_id"]}', json={
            'indicator_id': fixed.id, 'inputs': {'qualified': True}, 'note': '不应保存',
        })
        self.assertEqual(locked_edit.status_code, 409)

        self.client.post('/logout')
        self.client.post('/login', data={'username': 'admin', 'password': 'password123'})
        records_page = self.client.get('/admin/records')
        self.assertEqual(records_page.status_code, 200)
        self.assertIn('张老师'.encode(), records_page.data)
        response = self.client.post(
            f'/admin/records/{payload["record_id"]}/score',
            data={'final_score': '3.5', 'reason': '管理员复核'},
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        db.session.refresh(record)
        self.assertEqual(record.final_score, 3.5)
        self.assertTrue(record.admin_overridden)

        analytics = self.client.get('/api/admin/archive-analytics', query_string={
            'entity_type': 'teacher', 'entity_value': teacher.id,
        })
        self.assertEqual(analytics.status_code, 200)
        self.assertEqual(analytics.get_json()['summary']['records'], 2)
        archive_page = self.client.get('/admin/archive-analytics')
        self.assertIn('三年1班'.encode(), archive_page.data)
        self.assertIn(b'id="analytics-class" disabled', archive_page.data)
        self.assertIn('评价明细'.encode(), archive_page.data)
        class_analytics = self.client.get('/api/admin/archive-analytics', query_string={
            'entity_type': 'class', 'entity_value': '三年1班',
        })
        self.assertEqual(class_analytics.status_code, 200)
        self.assertEqual(class_analytics.get_json()['summary']['records'], 1)
        self.assertEqual(len(analytics.get_json()['details']), 2)

        no_ai = self.client.post('/api/admin/dictionary/from-document', data={
            'file': (BytesIO(b'# Test scheme'), 'scheme.md'),
        }, content_type='multipart/form-data')
        self.assertEqual(no_ai.status_code, 409)

        response = self.client.post('/admin/settings/scheme', data={
            'action': 'create', 'code': 'DEPT-002', 'name': '第二套考评方案',
            'academic_year': '2026—2027学年', 'description': '多方案测试',
        }, follow_redirects=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn('第二套考评方案'.encode(), response.data)
        second = EvaluationScheme.query.filter_by(code='DEPT-002').one()
        self.assertEqual(second.owner_user_id, None)
        second_year = AcademicYear.query.filter_by(scheme_id=second.id).one()
        self.assertEqual(second_year.label, '2026—2027学年')
        second_category = Category(
            academic_year_id=second_year.id, code='cross_scheme', name='跨方案测试', sort_order=1,
        )
        db.session.add(second_category)
        db.session.flush()
        second_indicator = Indicator(
            academic_year_id=second_year.id, category_id=second_category.id,
            code='cross_scheme_score', name='跨方案记录', scoring_type='fixed_score',
            scoring_rule_json='{"score": 4}', data_source='admin', secondary_tracking_json='{"enabled": false}',
        )
        db.session.add(second_indicator)
        db.session.flush()
        db.session.add(SchemeMembership(scheme_id=second.id, user_id=teacher.id, membership_role='participant'))
        db.session.add(EvaluationRecord(
            scheme_id=second.id, academic_year_id=second_year.id, indicator_id=second_indicator.id,
            target_user_id=teacher.id, submitted_by_user_id=User.query.filter_by(username='admin').one().id,
            source='admin', status='approved', input_json='{"qualified": true}', auto_score=4, final_score=4,
        ))
        db.session.commit()
        cross_scheme = self.client.get('/api/admin/archive-analytics', query_string={
            'entity_type': 'teacher', 'entity_value': teacher.id,
        }).get_json()
        self.assertEqual(cross_scheme['summary']['records'], 3)
        self.assertEqual(cross_scheme['summary']['schemes'], 2)
        self.assertTrue(any(row['scheme_code'] == 'DEPT-002' for row in cross_scheme['details']))


if __name__ == '__main__':
    unittest.main()
